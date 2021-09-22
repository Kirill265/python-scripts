import sys
import os
import calendar
import datetime
from datetime import timedelta
import xlsxwriter
import openpyxl
import pymysql
import psycopg2
from psycopg2.extras import DictCursor
from pymysql.cursors import DictCursor
import time
from keepass import key_pass
from win32com import client
import win32com

def report_generation(send_info):
    agent = send_info["agent"]
    sources_utm = send_info["sources_utm"]
    direction = send_info["direction"]
    date_from = send_info["date_from"]
    date_to = send_info["date_to"]
    msg_to_day = send_info["msg_to_day"]
    month = send_info["month"]
    sql_month = send_info["sql_month"]
    report_date = send_info["report_date"]
    SQL_DB = 'MySQL DB PROD'
    my_connection = pymysql.connect(
        host=key_pass(SQL_DB).url[:-5],
        port=int(key_pass(SQL_DB).url[-4:]),
        user=key_pass(SQL_DB).username,
        password=key_pass(SQL_DB).password,
        db='my',
        charset='utf8mb4',
        cursorclass=DictCursor
    )
    SQL_DB = 'PotgreSQL DB PROD'
    Postgre_connection = psycopg2.connect(
        host=key_pass(SQL_DB).url[:-5],
        port=int(key_pass(SQL_DB).url[-4:]),
        user=key_pass(SQL_DB).username,
        password=key_pass(SQL_DB).password,
        dbname='finance'
    )
    SQL_DB = 'PotgreSQL DB PROD'
    Postgre_connection_2 = psycopg2.connect(
        host=key_pass(SQL_DB).url[:-5],
        port=int(key_pass(SQL_DB).url[-4:]),
        user=key_pass(SQL_DB).username,
        password=key_pass(SQL_DB).password,
        dbname='mt5_report'
    )
    workbook_ = xlsxwriter.Workbook(direction+agent+' 01-'+msg_to_day+' '+month+' '+str(report_date.year)+'.xlsx')
    workbook_.formats[0].set_font_size(8.5)
    workbook_.formats[0].set_font_name('Tahoma')
    bold_blue_ = workbook_.add_format({'bold': True, 'fg_color': '#DDEBF7', 'align': 'center','valign': 'vcenter'})
    bold_blue_wrap = workbook_.add_format({'bold': True, 'fg_color': '#DDEBF7', 'align': 'center','valign': 'vcenter'})
    bold_blue_wrap.set_text_wrap()
    wrap_format = workbook_.add_format()
    wrap_format.set_text_wrap()
    wrap_format.set_font_size(8.5)
    wrap_format.set_font_name('Tahoma')
    number = workbook_.add_format({'num_format': '0.00','align': 'right'})
    rate = workbook_.add_format({'num_format': '0.0000','align': 'right'})
    minus = workbook_.add_format({'align': 'center','valign': 'vcenter'})
    percentage = workbook_.add_format({'num_format': '0.00%','align': 'center','valign': 'vcenter'})
    percentage.set_font_size(8.5)
    percentage.set_font_name('Tahoma')
    rub_reward = workbook_.add_format({'num_format': '0.00" ₽"'})
    usd_volume = workbook_.add_format({'num_format': '"$ "0.00'})
    worksheet_PL = workbook_.add_worksheet('PL')
    worksheet_PL.set_row(0, 15)
    worksheet_PL.write('A1', 'UTM', bold_blue_)
    worksheet_PL.set_column(0, 0, 15)
    worksheet_PL.write('B1', 'FIO', bold_blue_)
    worksheet_PL.set_column(1, 1, 25)
    worksheet_PL.write('C1', 'Phone number', bold_blue_)
    worksheet_PL.set_column(2, 2, 20)
    worksheet_PL.write('D1', 'LK', bold_blue_)
    worksheet_PL.write('E1', 'Login', bold_blue_)
    worksheet_PL.write('F1', 'Currency', bold_blue_)
    worksheet_PL.write('G1', 'Type', bold_blue_)
    worksheet_PL.set_column(3, 6, 15)
    worksheet_PL.write('H1', 'Created', bold_blue_)
    worksheet_PL.set_column(7, 7, 20)
    worksheet_PL.write('I1', 'Deposit', bold_blue_)
    worksheet_PL.write('J1', 'Withdrawal', bold_blue_)
    worksheet_PL.write('K1', 'Volume Lots', bold_blue_)
    worksheet_PL.set_column(8, 10, 15)
    worksheet_PL.write('L1', 'Convertation out', bold_blue_)
    worksheet_PL.write('M1', 'Convertation in', bold_blue_)
    worksheet_PL.set_column(11, 12, 18)
    worksheet_PL.write('N1', 'P/L (+Commission)', bold_blue_)
    worksheet_PL.set_column(13, 13, 20)
    worksheet_PL.write('O1', 'Equity '+str(report_date.day)+'.'+sql_month, bold_blue_)
    worksheet_PL.write('P1', 'Balance 1.'+sql_month, bold_blue_)
    worksheet_PL.write('Q1', 'Balance '+str(report_date.day)+'.'+sql_month, bold_blue_)
    worksheet_PL.write('R1', 'Rate 1.'+sql_month, bold_blue_)
    worksheet_PL.write('S1', 'Rate '+str(report_date.day)+'.'+sql_month, bold_blue_)
    worksheet_PL.set_column(14, 18, 15)
    worksheet_PL.write('T1', 'Equity '+str(report_date.day)+'.'+sql_month+', RUB', bold_blue_)
    worksheet_PL.write('U1', 'Balance 1.'+sql_month+', RUB', bold_blue_)
    worksheet_PL.write('V1', 'Balance '+str(report_date.day)+'.'+sql_month+', RUB', bold_blue_)
    worksheet_PL.set_column(19, 21, 20)
    worksheet_Reward = workbook_.add_worksheet('Reward')
    worksheet_Reward.set_row(0, 15)
    worksheet_Reward.write('A1', 'UTM', bold_blue_)
    worksheet_Reward.set_column(0, 0, 15)
    worksheet_Reward.write('B1', 'FIO', bold_blue_)
    worksheet_Reward.set_column(1, 1, 25)
    worksheet_Reward.write('C1', 'Phone number', bold_blue_)
    worksheet_Reward.set_column(2, 2, 20)
    worksheet_Reward.write('D1', 'LK', bold_blue_)
    worksheet_Reward.write('E1', 'Login', bold_blue_)
    worksheet_Reward.write('F1', 'Currency', bold_blue_)
    worksheet_Reward.write('G1', 'Volume Lots', bold_blue_)
    worksheet_Reward.write('H1', 'Volume USD', bold_blue_)
    worksheet_Reward.set_column(3, 7, 15)
    worksheet_Reward.write('I1', 'Reward, RUB', bold_blue_)
    worksheet_Reward.write('J1', 'Reward for convertation, RUB', bold_blue_wrap)
    worksheet_Reward.set_column(8, 9, 25)
    worksheet_Deals = workbook_.add_worksheet('Deals')
    worksheet_Deals.set_row(0, 15)
    worksheet_Deals.write('A1', 'UTM', bold_blue_)
    worksheet_Deals.set_column(0, 0, 15)
    worksheet_Deals.write('B1', 'FIO', bold_blue_)
    worksheet_Deals.set_column(1, 1, 25)
    worksheet_Deals.write('C1', 'Phone number', bold_blue_)
    worksheet_Deals.set_column(2, 2, 20)
    worksheet_Deals.write('D1', 'LK', bold_blue_)
    worksheet_Deals.write('E1', 'Login', bold_blue_)
    worksheet_Deals.write('F1', 'Currency', bold_blue_)
    worksheet_Deals.write('G1', 'Deal', bold_blue_)
    worksheet_Deals.write('H1', 'Order', bold_blue_)
    worksheet_Deals.set_column(3, 7, 15)
    worksheet_Deals.write('I1', 'Time', bold_blue_)
    worksheet_Deals.set_column(8, 8, 20)
    worksheet_Deals.write('J1', 'Action', bold_blue_)
    worksheet_Deals.write('K1', 'Entry', bold_blue_)
    worksheet_Deals.set_column(9, 10, 8.5)
    worksheet_Deals.write('L1', 'Symbol', bold_blue_)
    worksheet_Deals.write('M1', 'Volume', bold_blue_)
    worksheet_Deals.write('N1', 'Volume USD', bold_blue_)
    worksheet_Deals.write('O1', 'Price', bold_blue_)
    worksheet_Deals.write('P1', 'Reason', bold_blue_)
    worksheet_Deals.write('Q1', 'Dealer', bold_blue_)
    worksheet_Deals.write('R1', 'Profit', bold_blue_)
    worksheet_Deals.set_column(10, 17, 13)
    worksheet_Deals.write('S1', 'Rate', bold_blue_)
    worksheet_Deals.write('T1', 'Profit, RUB', bold_blue_)
    worksheet_Deals.set_column(18, 19, 15)
    worksheet_DW = workbook_.add_worksheet('DepositWithdrawal')
    worksheet_DW.set_row(0, 15)
    worksheet_DW.write('A1', 'UTM', bold_blue_)
    worksheet_DW.set_column(0, 0, 15)
    worksheet_DW.write('B1', 'FIO', bold_blue_)
    worksheet_DW.set_column(1, 1, 15)
    worksheet_DW.write('C1', 'Phone number', bold_blue_)
    worksheet_DW.set_column(2, 2, 20)
    worksheet_DW.write('D1', 'LK', bold_blue_)
    worksheet_DW.write('E1', 'Login', bold_blue_)
    worksheet_DW.write('F1', 'Currency', bold_blue_)
    worksheet_DW.set_column(3, 5, 15)
    worksheet_DW.write('G1', 'Time', bold_blue_)
    worksheet_DW.write('H1', 'Balance before', bold_blue_)
    worksheet_DW.set_column(6, 7, 20)
    worksheet_DW.write('I1', 'Deposit / Withdrawal', bold_blue_)
    worksheet_DW.set_column(8, 8, 25)
    worksheet_DW.write('J1', 'Rate', bold_blue_)
    worksheet_DW.set_column(9, 9, 15)
    worksheet_DW.write('K1', 'Balance before, RUB', bold_blue_)
    worksheet_DW.set_column(10, 10, 25)
    worksheet_DW.write('L1', 'Deposit / Withdrawal, RUB', bold_blue_)
    worksheet_DW.set_column(11, 11, 30)
    worksheet_ML = workbook_.add_worksheet('Margin Level')
    worksheet_ML.set_row(0, 15)
    worksheet_ML.write('A1', 'UTM', bold_blue_)
    worksheet_ML.set_column(0, 0, 15)
    worksheet_ML.write('B1', 'FIO', bold_blue_)
    worksheet_ML.set_column(1, 1, 25)
    worksheet_ML.write('C1', 'Phone number', bold_blue_)
    worksheet_ML.set_column(2, 2, 20)
    worksheet_ML.write('D1', 'LK', bold_blue_)
    worksheet_ML.write('E1', 'Login', bold_blue_)
    worksheet_ML.write('F1', 'Currency', bold_blue_)
    worksheet_ML.set_column(3, 5, 15)
    worksheet_ML.write('G1', 'Margin Level '+str(report_date.day)+'.'+sql_month, bold_blue_)
    worksheet_ML.set_column(6, 6, 20)
    workbook_sum = xlsxwriter.Workbook(direction+agent.lower()+' рассчёт 01-'+msg_to_day+' '+month+' '+str(report_date.year)+'.xlsx')
    border_sum = workbook_sum.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
    rub = workbook_sum.add_format({'num_format': '0.00"₽"'})
    usd = workbook_sum.add_format({'num_format': '"$"0.00'})
    rub_border = workbook_sum.add_format({'num_format': '0.00"₽"','border': 1,'align': 'center','valign': 'vcenter'})
    usd_border = workbook_sum.add_format({'num_format': '"$"0.00','border': 1,'align': 'center','valign': 'vcenter'})
    bold = workbook_sum.add_format({'bold': True})
    bold_border = workbook_sum.add_format({'bold': True,'border': 1,'align': 'center','valign': 'vcenter'})
    workbook_sum.formats[0].set_font_size(8.5)
    workbook_sum.formats[0].set_font_name('Tahoma')
    worksheet_sum = workbook_sum.add_worksheet('Итог')
    worksheet_sum.set_default_row(12)
    worksheet_sum.set_row(0, 15)
    worksheet_sum.write('A1', 'Агент', bold_border)
    worksheet_sum.write('B1', 'Клиент', bold_border)
    worksheet_sum.set_column(0, 1, 15)
    worksheet_sum.write('C1', 'Сумма за клиента', bold_border)
    worksheet_sum.set_column(2, 2, 25)
    worksheet_sum.write('D1', 'Итоговая сумма', bold_border)
    worksheet_sum.set_column(3, 3, 20)
    with my_connection.cursor() as cursor:
        query = """
                SELECT 
                CONCAT("'",c.universal_id,"',") AS UID_for_Postgre 
                FROM customer_utm cu
                LEFT JOIN customer c ON cu.customer_id = c.id
                LEFT JOIN utm u ON cu.utm_id = u.id
                WHERE u.utm_source IN ("""+sources_utm+""")
         """
        cursor.execute(query)
        UIDS = cursor.fetchall()
        customer_id = ''
        for UID in UIDS:
            customer_id += UID["UID_for_Postgre"]
        query = """
                SELECT CONCAT(a.login,',') AS login_comma
                FROM account a
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id 
                WHERE u.utm_source IN ("""+sources_utm+""")
                ORDER BY Login
         """
        cursor.execute(query)
        logins_comma = cursor.fetchall()
        login_for_mt5 = ''
        for login_comma in logins_comma:
            login_for_mt5 += login_comma["login_comma"]
        query = """
                SELECT CONCAT("'",a.login,"',") AS login_str_comma
                FROM account a
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id 
                WHERE u.utm_source IN ("""+sources_utm+""")
                ORDER BY Login
         """
        cursor.execute(query)
        logins_str_comma = cursor.fetchall()
        login_for_conv = ''
        for login_str_comma in logins_str_comma:
            login_for_conv += login_str_comma["login_str_comma"]
        query = """
                SET @@time_zone = \"+3:00\";
        """
        cursor.execute(query)
        query = """
                SELECT DATE(crh.date) as currency_date, crh.value FROM currency_rate_history crh
                WHERE crh.from_id = 1
                AND crh.to_id = 3
                AND crh.date BETWEEN DATE(\""""+date_from+"""\") AND DATE(\""""+date_to+"""\")
                ORDER BY crh.date
        """
        cursor.execute(query)
        currency_rates = cursor.fetchall()
        query = """
                SELECT DISTINCT u.utm_source FROM platform_mt5_deal pmd
                LEFT JOIN account a ON pmd.login = a.login
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id
                WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                and u.utm_source IN ("""+sources_utm+""")
        """
        cursor.execute(query)
        utm_sources = cursor.fetchall()
        k = 1
        done_counter = 0
        login_list = []
        for utm_source in utm_sources:
            query = """
                    SELECT a.login FROM account a
                    LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                    LEFT JOIN utm u ON cu.utm_id = u.id
                    WHERE u.utm_source = '"""+utm_source["utm_source"]+"""'
                    AND a.login IN 
                    (SELECT DISTINCT pmd.login FROM platform_mt5_deal pmd
                    WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\")
            """
            cursor.execute(query)
            logins = cursor.fetchall()
            i = 1
            for login in logins:
                i += 1
                k += 1
                login_list.append(str(login["login"]))
                worksheet_login = workbook_sum.add_worksheet(str(login["login"]))
                worksheet_login.set_default_row(12)
                worksheet_login.set_row(0, 20)
                worksheet_login.write('A1', 'id', bold)
                worksheet_login.write('B1', 'login', bold)
                worksheet_login.set_column(0, 1, 13)
                worksheet_login.write('C1', 'mt5_order_id', bold)
                worksheet_login.set_column(2, 2, 15)
                worksheet_login.write('D1', 'action', bold)
                worksheet_login.write('E1', 'entry', bold)
                worksheet_login.set_column(3, 4, 7)
                worksheet_login.write('F1', 'symbol', bold)
                worksheet_login.set_column(5, 5, 12)
                worksheet_login.write('G1', 'created_at', bold)
                worksheet_login.write('H1', 'date', bold)
                worksheet_login.set_column(6, 7, 18)
                worksheet_login.write('I1', 'price', bold)
                worksheet_login.set_column(8, 8, 8)
                worksheet_login.write('J1', 'volume', bold)
                worksheet_login.set_column(9, 9, 9)
                worksheet_login.write('K1', 'profit', bold)
                worksheet_login.set_column(10, 10, 10) 
                worksheet_login.write('L1', 'volume_usd', bold)
                worksheet_login.set_column(11, 11, 14)
                worksheet_login.write('M1', 'expert_position_id', bold)
                worksheet_login.set_column(12, 12, 20)
                worksheet_login.write('N1', 'Одна поза', bold)
                worksheet_login.set_column(13, 13, 11)
                worksheet_login.write('O1', 'Длительность позиции', bold)
                worksheet_login.set_column(14, 14, 15)
                worksheet_login.write('P1', 'Короткая', bold)
                worksheet_login.set_column(15, 15, 10)
                worksheet_login.write('Q1', 'Вознаграждение, USD', bold)
                worksheet_login.set_column(16, 16, 18)
                worksheet_login.write('R1', 'Курс USDRUR на дату сделки', bold)
                worksheet_login.set_column(17, 17, 14)
                worksheet_login.write('S1', 'Вознаграждение, RUR', bold)
                worksheet_login.set_column(18, 18, 18)
                worksheet_login.write('V1', 'USD без округления', bold)
                worksheet_login.write('W1', 'RUB без округления', bold)
                worksheet_login.set_column(21, 22, 12)
                worksheet_login.write('U2', 'Вознаграждение без учета коротких сделок')
                worksheet_login.write('U3', 'Вознаграждение с вычетом коротких сделок')
                worksheet_login.set_column(20, 20, 23)
                worksheet_login.write('V2', '=SUM(L:L)/1000000*20', usd)
                worksheet_login.write('V3', '=SUM(Q:Q)', usd)
                worksheet_login.write('W3', '=SUM(S:S)', rub)
                query = """
                        SELECT
                        pmd.id
                        , pmd.login
                        , pmd.mt5_order_id
                        , pmd.action
                        , pmd.entry
                        , pmd.symbol
                        , pmd.created_at
                        , date(pmd.created_at) AS date
                        , pmd.price
                        , pmd.volume
                        , pmd.profit
                        , pmd.volume_usd
                        , pmd.expert_position_id
                        FROM platform_mt5_deal pmd
                        LEFT JOIN account a ON pmd.login = a.login
                        LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                        LEFT JOIN utm u ON cu.utm_id = u.id
                        LEFT JOIN account_group ag ON a.group_id = ag.id
                        WHERE a.login = '"""+str(login["login"])+"""'
                        AND
                        pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                        ORDER BY pmd.id
                """
                cursor.execute(query)
                deals = cursor.fetchall()
                j = 1
                for deal in deals:
                    j += 1
                    worksheet_login.write(f'A{j}', deal["id"])
                    worksheet_login.write(f'B{j}', deal["login"])
                    worksheet_login.write(f'C{j}', deal["mt5_order_id"])
                    worksheet_login.write(f'D{j}', deal["action"])
                    worksheet_login.write(f'E{j}', deal["entry"])
                    worksheet_login.write(f'F{j}', deal["symbol"])
                    worksheet_login.write(f'G{j}', str(deal["created_at"]))
                    worksheet_login.write(f'H{j}', str(deal["date"]))
                    worksheet_login.write(f'I{j}', deal["price"])
                    worksheet_login.write(f'J{j}', deal["volume"])
                    worksheet_login.write(f'K{j}', deal["profit"])
                    worksheet_login.write(f'L{j}', deal["volume_usd"])
                    worksheet_login.write(f'M{j}', deal["expert_position_id"])
                    worksheet_login.write(f'N{j}', '=M'+str(j)+'=M'+str(j-1))
                    worksheet_login.write(f'O{j}', '=IF(N'+str(j)+' = TRUE,(G'+str(j)+'-G'+str(j-1)+')*24*60*60,"> 10")')
                    worksheet_login.write(f'P{j}', '=IF(O'+str(j)+'=">10",FALSE,AND(O'+str(j)+'<10,E'+str(j-1)+'<>1,E'+str(j)+'<>0))')
                    worksheet_login.write(f'Q{j}', '=IF(P'+str(j)+' = TRUE,0,L'+str(j)+'/1000000*20)')
                    worksheet_login.write(f'R{j}', '=VLOOKUP(H'+str(j)+',\'Курс ЦБ\'!A:B,2,FALSE)')
                    worksheet_login.write(f'S{j}', '=Q'+str(j)+'*R'+str(j)+'')
                worksheet_sum.write(f'B{k}',str(login["login"]), border_sum)
                worksheet_sum.write(f'C{k}','=ROUND(\''+worksheet_login.name+'\'!$W$3,2)', rub_border)
            if i == 2:
                worksheet_sum.write(f'D{k}','=C'+str(k), rub_border)
                worksheet_sum.write(f'A{k}',str(utm_source["utm_source"]), border_sum)
            else:
                worksheet_sum.merge_range(f'D{k-i+2}:D{k}','=SUM(C'+str(k-i+2)+':C'+str(k)+')', rub_border)
                worksheet_sum.merge_range(f'A{k-i+2}:A{k}',str(utm_source["utm_source"]), border_sum)
            done_counter += 1
        worksheet_currency = workbook_sum.add_worksheet('Курс ЦБ')
        worksheet_currency.set_default_row(12)
        i = 0
        for currency_rate in currency_rates:
            i += 1
            worksheet_currency.write(f'A{i}',str(currency_rate["currency_date"]))
            worksheet_currency.set_column(0, 0, 12)
            worksheet_currency.write(f'B{i}',currency_rate["value"])
            worksheet_currency.set_column(1, 1, 8)
        query = """
                SELECT UPPER(c.name) AS Currency, crh.value, crh.date FROM currency_rate_history crh
                LEFT JOIN currency c ON crh.from_id = c.id
                WHERE crh.date BETWEEN date(\""""+date_from+"""\") AND date(\""""+date_to+"""\")
                AND crh.to_id = 3
                AND c.name IN ('eur', 'usd')
         """
        cursor.execute(query)
        currencys = cursor.fetchall()
        currency_dict = {"USD":{}, "EUR":{}}
        for currency in currencys:
            currency_dict[currency["Currency"]][currency["date"]] = currency["value"]
    with Postgre_connection.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:    
        query = """
                SELECT a.login, 
                COALESCE(conv_from.amount, 0.00) as volume_out
                ,COALESCE(conv_to.amount, 0.00) AS volume_in
                FROM account a
                LEFT JOIN 
                (SELECT
                a_from.login AS login
                , SUM(ROUND(- c.amount_from / 100.0, 2)) AS amount
                FROM convertation c
                LEFT JOIN account a_from ON c.account_id_from = a_from.id
                WHERE c.created_at BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                AND c.status = 3
                GROUP BY a_from.login
                ) AS conv_from ON conv_from.login = a.login
                LEFT JOIN
                (SELECT
                a_to.login AS login
                , SUM(ROUND(c.amount_to / 100.0, 2)) AS amount
                FROM convertation c
                LEFT JOIN account a_to ON c.account_id_to = a_to.id
                WHERE c.created_at BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                AND c.status = 3
                GROUP BY a_to.login
                ) AS conv_to ON conv_to.login = a.login
                WHERE (conv_from.amount IS NOT NULL
                OR conv_to.amount IS NOT NULL
                )
                AND a.customer_id IN (
                -- id клиентов
                """+customer_id[:-1]+"""
                ) 
                ORDER BY a.login
        """
        cursor.execute(query)
        convertations = cursor.fetchall()
        convertation_dict = {}
        for convertation in convertations:
            convertation_dict[str(convertation["login"])] = {"out":convertation["volume_out"], "in":convertation["volume_in"], "reward":0.00}
        if login_for_conv[:-1] != '':
            query = """
                    SELECT
                    a_to.login AS login	 
                    , CASE
                    WHEN c.operation_type = 'BUY' THEN ROUND(c.amount_to / 100.0, 2) - ROUND(c.amount_from / c.market_rate / 100.0, 2)
                    WHEN c.operation_type = 'SELL' THEN ROUND(c.amount_to / 100.0, 2) - ROUND(c.amount_from * c.market_rate / 100.0, 2)
                    END AS finrez
                    , DATE(c.created_at) as conv_date
                    , a_to.currency AS currency
                    FROM convertation c
                    LEFT JOIN account a_to ON c.account_id_to = a_to.id
                    WHERE c.created_at BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                    AND c.status = 3
                    AND a_to.login IN (
                    """+login_for_conv[:-1]+"""
                    ) 
                    ORDER BY a_to.login;
            """
            cursor.execute(query)
            convert_reward = cursor.fetchall()
            for conv_reward in convert_reward:
                if conv_reward["currency"] == 'RUB':
                    convertation_dict[str(conv_reward["login"])]["reward"] += -round(float(conv_reward["finrez"])/2,2)
                else:
                    convertation_dict[str(conv_reward["login"])]["reward"] += -round(float(conv_reward["finrez"])*float(currency_dict[conv_reward["currency"]][conv_reward["conv_date"]])/2,2)
    Postgre_connection.close()
    workbook_sum.close()
    xl = win32com.client.DispatchEx('Excel.Application')
    xl.Visible = False
    wb = xl.Workbooks.Open(direction+agent+" рассчёт 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx")
    wb.Close(True)
    xl.Quit()
    with Postgre_connection_2.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:    
        if login_for_mt5[:-1] != '':
            query = """
                    SELECT mt5a."Login"
                    , ROUND(COALESCE(Dps.Deposit,0)::NUMERIC,2) AS Deposit
                    , ROUND(COALESCE(-Wth.Withdrawal,0)::NUMERIC,2) AS Withdrawal
                    , ROUND(COALESCE(PL.profit,0)::NUMERIC,2) AS Profit
                    FROM mt5_accounts mt5a
                    LEFT JOIN (
                    SELECT "Login", SUM("Profit") AS profit FROM mt5_deals
                    WHERE "Action" IN (0,1,7)
                    AND "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                    GROUP BY "Login") AS PL ON mt5a."Login" = PL."Login"
                    LEFT JOIN (
                    SELECT "Login", SUM("Profit") AS Deposit FROM mt5_deals
                    WHERE "Action" = 2
                    AND  "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                    AND (
                    "Comment" = ''
                    OR
                    "Comment" LIKE '%Deposit%'
                    OR
                    "Comment" LIKE '%Возврат%'
                    OR
                    "Comment" LIKE '%Refund%'
                    )
                    GROUP BY "Login") AS Dps ON Dps."Login" = mt5a."Login"
                    LEFT JOIN (
                    SELECT "Login", SUM("Profit") AS Withdrawal FROM mt5_deals
                    WHERE "Action" = 2
                    AND  "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                    AND (
                    "Comment" LIKE '%Withdrawal%'
                    OR
                    "Comment" LIKE '%Удержание%'
                    OR
                    "Comment" LIKE '%удержание%'
                    )
                    GROUP BY "Login") AS Wth ON Wth."Login" = mt5a."Login" 
                    WHERE mt5a."Login" IN (
                    """+login_for_mt5[:-1]+"""
                    );
            """
            cursor.execute(query)
            OPRDS_plus_PL = cursor.fetchall()
            OPRDS_PL_dict = {}
            for OPRDS_PL in OPRDS_plus_PL:
                OPRDS_PL_dict[str(OPRDS_PL["Login"])] = {"deposit":OPRDS_PL["deposit"], "withdrawal":OPRDS_PL["withdrawal"], "profit":OPRDS_PL["profit"]}
            query = """
                    SELECT "md_over"."Login", mu."State", mu."FirstName"
                    , ROUND(("md_over"."Balance" - "md_over"."Profit")::numeric, 2) AS "Balance_before"
                    , "md_over"."Profit", "md_over"."TimeMsc", DATE("md_over"."TimeMsc") AS date_dw
                    FROM (
                    SELECT md."Login", md."Profit", md."Comment", md."TimeMsc", md."Action"
                    , SUM(md."Profit") OVER(PARTITION BY md."Login" ORDER BY md."TimeMsc") AS "Balance"
                    FROM mt5_deals md
                    ORDER BY md."Login"
                    ) AS "md_over"
                    LEFT JOIN mt5_users mu ON "md_over"."Login" = mu."Login"
                    WHERE "md_over"."Action" = 2
                    AND  "md_over"."TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
                    AND (
                    "md_over"."Comment" = ''
                    OR
                    "md_over"."Comment" LIKE '%Deposit%'
                    OR
                    "md_over"."Comment" LIKE '%Возврат%'
                    OR
                    "md_over"."Comment" LIKE '%Refund%'
                    OR
                    "md_over"."Comment" LIKE '%Withdrawal%'
                    OR
                    "md_over"."Comment" LIKE '%Удержание%'
                    OR
                    "md_over"."Comment" LIKE '%удержание%'
                    )
                    AND "md_over"."Login" IN (
                    """+login_for_mt5[:-1]+"""
                    );
            """
            cursor.execute(query)
            DepositAndWithdrawal = cursor.fetchall()
        Postgre_connection_2.close()
    with my_connection.cursor() as cursor:
        query = """
                SET @@time_zone = "+3:00";
         """
        cursor.execute(query)
        query = """
                SELECT a.login AS 'Login'
                , a.created_at AS 'Create'
                , CONCAT(ci.last_name_ru,' ',ci.first_name_ru,' ',ci.middle_name_ru) AS 'FIO'
                , cus.mobile_phone AS 'Phone'
                , u.utm_source AS 'UTM'
                , a.customer_id AS 'LK'
                , CASE
                    WHEN ag.name LIKE '%Hedge%' THEN 'hedge'
                    ELSE 'netting'
                    END AS 'Type'
                , UPPER(c.name) AS 'Currency'
                , IFNULL(ROUND(pmd.volume_sum, 2), 0) AS 'Volume_Lots'
                , IFNULL(ROUND(EquityTo.Equity,2),0) AS Equity_to
                , IFNULL(ROUND(EquityFrom.Balance,2),0) AS Balance_from
                , IFNULL(ROUND(EquityTo.Balance,2),0) AS Balance_to
                , ROUND(currency_convert(a.currency_id, 3, DATE(\""""+date_from+"""\")),4) AS CBR_from
                , ROUND(currency_convert(a.currency_id, 3, DATE(\""""+date_to+"""\")),4) AS CBR_to
                , ROUND(IFNULL(ROUND(EquityTo.Equity,2),0)*currency_convert(a.currency_id, 3, DATE(\""""+date_to+"""\")),2) AS Equity_rub_to
                , ROUND(IFNULL(ROUND(EquityFrom.Balance,2),0)*currency_convert(a.currency_id, 3, DATE(\""""+date_from+"""\")),2) AS Balance_rub_from
                , ROUND(IFNULL(ROUND(EquityTo.Balance,2),0)*currency_convert(a.currency_id, 3, DATE(\""""+date_to+"""\")),2) AS Balance_rub_to
                , IFNULL(ROUND(EquityTo.MarginLevel/100,4),0) AS Margin_Level
                FROM account a
                LEFT JOIN currency c ON a.currency_id = c.id
                LEFT JOIN customer cus ON a.customer_id = cus.id
                LEFT JOIN customer_individual ci ON a.customer_id = ci.customer_id
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id
                LEFT JOIN account_group ag ON ag.id = a.group_id
                LEFT JOIN
                (
                SELECT platform_mt5_deal.login
                , SUM(platform_mt5_deal.volume) AS volume_sum
                FROM platform_mt5_deal 
                WHERE platform_mt5_deal.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                GROUP BY platform_mt5_deal.login
                ) AS pmd ON pmd.login = a.login
                  LEFT JOIN
                (
                SELECT
                m5ad1.login
                , m5ad1.MarginLevel
                , m5ad1.Balance
                , m5ad1.Equity
                FROM mt5_events.mt5accountdaily m5ad1
                WHERE DATE(DATE_ADD("1970-01-01 00:00:00", INTERVAL ROUND(m5ad1.TimestampMS/1000 - 3600, 0) SECOND)) = DATE(\""""+date_to+"""\")
                ) AS EquityTo ON a.login = EquityTo.login
                LEFT JOIN
                (
                SELECT
                m5ad2.login
                , m5ad2.Balance
                FROM mt5_events.mt5accountdaily m5ad2
                WHERE DATE(DATE_ADD("1970-01-01 00:00:00", INTERVAL ROUND(m5ad2.TimestampMS/1000 - 3600, 0) SECOND)) = DATE(DATE_ADD(\""""+date_from+"""\", INTERVAL -1 DAY))
                ) AS EquityFrom ON a.login = EquityFrom.login
                WHERE u.utm_source IN ("""+sources_utm+""")
                ORDER BY Login
         """
        cursor.execute(query)
        PL_all = cursor.fetchall()
        wb = openpyxl.load_workbook(direction+agent+" рассчёт 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx",data_only=True)
        Login_utm_dict = {}
        j = 1
        for PL_one in PL_all:
            j += 1
            Login_utm_dict[str(PL_one["Login"])] = {"utm":PL_one["UTM"],"fio":PL_one["FIO"],"phone":PL_one["Phone"],"lk":PL_one["LK"],"currency":PL_one["Currency"]}
            worksheet_PL.write(f'A{j}', PL_one["UTM"])
            worksheet_PL.write(f'B{j}', PL_one["FIO"],wrap_format)
            worksheet_PL.write(f'C{j}', PL_one["Phone"])
            worksheet_PL.write(f'D{j}', PL_one["LK"])
            worksheet_PL.write(f'E{j}', PL_one["Login"])
            worksheet_PL.write(f'F{j}', PL_one["Currency"])
            worksheet_PL.write(f'G{j}', PL_one["Type"])
            worksheet_PL.write(f'H{j}', str(PL_one["Create"]))
            worksheet_PL.write(f'I{j}', OPRDS_PL_dict[str(PL_one["Login"])]["deposit"], number)
            worksheet_PL.write(f'J{j}', OPRDS_PL_dict[str(PL_one["Login"])]["withdrawal"], number)
            worksheet_PL.write(f'K{j}', PL_one["Volume_Lots"], number)
            try:
                worksheet_PL.write(f'L{j}', convertation_dict[str(PL_one["Login"])]["out"], number)
                worksheet_PL.write(f'M{j}', convertation_dict[str(PL_one["Login"])]["in"], number)
            except KeyError:
                worksheet_PL.write(f'L{j}', 0.00, number)
                worksheet_PL.write(f'M{j}', 0.00, number)
            worksheet_PL.write(f'N{j}', OPRDS_PL_dict[str(PL_one["Login"])]["profit"], number)
            worksheet_PL.write(f'O{j}', PL_one["Equity_to"], number)
            worksheet_PL.write(f'P{j}', PL_one["Balance_from"], number)
            worksheet_PL.write(f'Q{j}', PL_one["Balance_to"], number)
            worksheet_PL.write(f'R{j}', PL_one["CBR_from"], rate)
            worksheet_PL.write(f'S{j}', PL_one["CBR_to"], rate)
            worksheet_PL.write(f'T{j}', PL_one["Equity_rub_to"], number)
            worksheet_PL.write(f'U{j}', PL_one["Balance_rub_from"], number)
            worksheet_PL.write(f'V{j}', PL_one["Balance_rub_to"], number)
            worksheet_Reward.write(f'A{j}', PL_one["UTM"])
            worksheet_Reward.write(f'B{j}', PL_one["FIO"],wrap_format)
            worksheet_Reward.write(f'C{j}', PL_one["Phone"])
            worksheet_Reward.write(f'D{j}', PL_one["LK"])
            worksheet_Reward.write(f'E{j}', PL_one["Login"])
            worksheet_Reward.write(f'F{j}', PL_one["Currency"])
            worksheet_Reward.write(f'G{j}', PL_one["Volume_Lots"], number)
            if str(PL_one["Login"]) in login_list:
                sheet = wb[str(PL_one["Login"])]
                RewardSelected = sheet.cell(row = 3, column = 22).value
                in_USD = round(RewardSelected*1000000/20,2)
                worksheet_Reward.write(f'H{j}', '='+str(in_USD), number)
                RewardSelected = sheet.cell(row = 3, column = 23).value
                in_RUB = round(RewardSelected,2)
                worksheet_Reward.write(f'I{j}', '='+str(in_RUB), number)
            else:
                worksheet_Reward.write(f'H{j}', 0.00, number)
                worksheet_Reward.write(f'I{j}', 0.00, number)
            try:
                worksheet_Reward.write(f'J{j}', convertation_dict[str(PL_one["Login"])]["reward"], number)
            except KeyError:
                worksheet_Reward.write(f'J{j}', 0.00, number)
            worksheet_ML.write(f'A{j}', PL_one["UTM"])
            worksheet_ML.write(f'B{j}', PL_one["FIO"],wrap_format)
            worksheet_ML.write(f'C{j}', PL_one["Phone"])
            worksheet_ML.write(f'D{j}', PL_one["LK"])
            worksheet_ML.write(f'E{j}', PL_one["Login"])
            worksheet_ML.write(f'F{j}', PL_one["Currency"])
            worksheet_ML.write(f'G{j}', PL_one["Margin_Level"],percentage)
        query = """
                SELECT c.created_at AS 'Create'
                , CONCAT(ci.last_name_ru,' ',ci.first_name_ru,' ',ci.middle_name_ru) AS 'FIO'
                , c.mobile_phone AS 'Phone'
                , u.utm_source AS 'UTM'
                , c.id AS 'LK'
                FROM customer_utm cu
                LEFT JOIN utm u ON cu.utm_id = u.id
                LEFT JOIN customer c ON cu.customer_id = c.id
                LEFT JOIN customer_individual ci ON c.id = ci.customer_id
                WHERE u.utm_source IN ("""+sources_utm+""")
                AND c.id NOT IN (
                SELECT distinct a.customer_id FROM account a
                )
         """
        cursor.execute(query)
        customerID_all = cursor.fetchall()
        for customerID in customerID_all:
            j += 1
            worksheet_PL.write(f'A{j}', customerID["UTM"])
            worksheet_PL.write(f'B{j}', customerID["FIO"],wrap_format)
            worksheet_PL.write(f'C{j}', customerID["Phone"])
            worksheet_PL.write(f'D{j}', customerID["LK"])
            worksheet_PL.write(f'E{j}', '-',minus)
            worksheet_PL.write(f'F{j}', '-',minus)
            worksheet_PL.write(f'G{j}', '-',minus)
            #worksheet_PL.write(f'H{j}', '-',minus)
            worksheet_PL.write(f'H{j}', str(customerID["Create"]))
            worksheet_PL.write(f'I{j}', '-',minus)
            worksheet_PL.write(f'J{j}', '-',minus)
            worksheet_PL.write(f'K{j}', '-',minus)
            worksheet_PL.write(f'L{j}', '-',minus)
            worksheet_PL.write(f'M{j}', '-',minus)
            worksheet_PL.write(f'N{j}', '-',minus)
            worksheet_PL.write(f'O{j}', '-',minus)
            worksheet_PL.write(f'P{j}', '-',minus)
            worksheet_PL.write(f'Q{j}', '-',minus)
            worksheet_PL.write(f'R{j}', '-',minus)
            worksheet_PL.write(f'S{j}', '-',minus)
            worksheet_PL.write(f'T{j}', '-',minus)
            worksheet_PL.write(f'U{j}', '-',minus)
            worksheet_PL.write(f'V{j}', '-',minus)
            worksheet_Reward.write(f'A{j}', customerID["UTM"])
            worksheet_Reward.write(f'B{j}', customerID["FIO"],wrap_format)
            worksheet_Reward.write(f'C{j}', customerID["Phone"])
            worksheet_Reward.write(f'D{j}', customerID["LK"])
            worksheet_Reward.write(f'E{j}', '-',minus)
            worksheet_Reward.write(f'F{j}', '-',minus)
            worksheet_Reward.write(f'G{j}', '-',minus)
            worksheet_Reward.write(f'H{j}', '-',minus)
            worksheet_Reward.write(f'I{j}', '-',minus)
            worksheet_Reward.write(f'J{j}', '-',minus)
            worksheet_ML.write(f'A{j}', customerID["UTM"])
            worksheet_ML.write(f'B{j}', customerID["FIO"],wrap_format)
            worksheet_ML.write(f'C{j}', customerID["Phone"])
            worksheet_ML.write(f'D{j}', customerID["LK"])
            worksheet_ML.write(f'E{j}', '-',minus)
            worksheet_ML.write(f'F{j}', '-',minus)
            worksheet_ML.write(f'G{j}', '-',minus)
        s = 1
        if login_for_mt5[:-1] != '':
            for DandW in DepositAndWithdrawal:
                s += 1
                worksheet_DW.write(f'A{s}', Login_utm_dict[str(DandW["Login"])]["utm"])
                worksheet_DW.write(f'B{s}', Login_utm_dict[str(DandW["Login"])]["fio"],wrap_format)
                worksheet_DW.write(f'C{s}', Login_utm_dict[str(DandW["Login"])]["phone"])
                worksheet_DW.write(f'D{s}', Login_utm_dict[str(DandW["Login"])]["lk"])
                worksheet_DW.write(f'E{s}', DandW["Login"])
                worksheet_DW.write(f'F{s}', Login_utm_dict[str(DandW["Login"])]["currency"])
                worksheet_DW.write(f'G{s}', str(str(DandW["TimeMsc"]).split('.')[0]))
                worksheet_DW.write(f'H{s}', DandW["Balance_before"],number)
                worksheet_DW.write(f'I{s}', DandW["Profit"],number)
                if Login_utm_dict[str(DandW["Login"])]["currency"] == 'RUB':
                    worksheet_DW.write(f'J{s}', round(float(1),4))
                    worksheet_DW.write(f'K{s}', round(float(DandW["Balance_before"]),2),number)
                    worksheet_DW.write(f'L{s}', round(float(DandW["Profit"]),2),number)
                else:
                    worksheet_DW.write(f'J{s}', round(float(currency_dict[Login_utm_dict[str(DandW["Login"])]["currency"]][DandW["date_dw"]]),4),rate)
                    worksheet_DW.write(f'K{s}', round(float(DandW["Balance_before"])*float(currency_dict[Login_utm_dict[str(DandW["Login"])]["currency"]][DandW["date_dw"]]),2),number)
                    worksheet_DW.write(f'L{s}', round(float(DandW["Profit"])*float(currency_dict[Login_utm_dict[str(DandW["Login"])]["currency"]][DandW["date_dw"]]),2),number)
        query = """
                SELECT
                CONCAT(ci.last_name_ru,' ',ci.first_name_ru,' ',ci.middle_name_ru) AS 'FIO'
                , cus.mobile_phone AS 'Phone'
                , a.customer_id AS 'LK'
                , pmd.id AS 'Deal'
                , pmd.mt5_order_id AS 'Order'
                , pmd.login AS 'Login'
                , u.utm_source AS 'UTM'
                , pmd.created_at AS 'Time'
                , IF(pmd.action = 0, 'buy', 'sell') AS 'Action'
                , IF(pmd.entry = 0, 'in', IF(pmd.entry = 1, 'out', 'in/out')) AS 'Entry'
                , pmd.symbol AS 'Symbol'
                , pmd.volume AS 'Volume'
                , pmd.volume_usd AS 'Volume_USD'
                , pmd.price AS 'Price'
                , CASE
                WHEN pmd.reason = 0 THEN 'Client'
                WHEN pmd.reason = 1 THEN 'Expert'
                WHEN pmd.reason = 2 THEN 'Dealer'
                WHEN pmd.reason = 3 THEN 'Stop loss'
                WHEN pmd.reason = 4 THEN 'Take profit'
                WHEN pmd.reason = 5 THEN 'Stop out'
                WHEN pmd.reason = 16 THEN 'Mobile'
                WHEN pmd.reason = 17 THEN 'Web'
                ELSE pmd.reason
                END AS 'Reason'
                , pmd.profit AS 'Profit'
                , ROUND(currency_convert(a.currency_id, 3, DATE(pmd.created_at)),4) AS 'CBR'
                , ROUND(pmd.profit*currency_convert(a.currency_id, 3, DATE(pmd.created_at)),2) AS 'Result'
                , pmd.dealer_id AS 'Dealer'
                , UPPER(c.name) AS 'Currency'
                FROM platform_mt5_deal pmd
                LEFT JOIN account a ON pmd.login = a.login
                LEFT JOIN customer cus ON a.customer_id = cus.id
                LEFT JOIN customer_individual ci ON a.customer_id = ci.customer_id
                LEFT JOIN currency c ON a.currency_id = c.id
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id
                WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                AND (
                u.utm_source IN ("""+sources_utm+""")
                )
                ORDER BY Deal
        """
        cursor.execute(query)
        Deals = cursor.fetchall()
        m = 1
        for Deal in Deals:
            m += 1
            worksheet_Deals.write(f'A{m}', Deal["UTM"])
            worksheet_Deals.write(f'B{m}', Deal["FIO"],wrap_format)
            worksheet_Deals.write(f'C{m}', Deal["Phone"])
            worksheet_Deals.write(f'D{m}', Deal["LK"])
            worksheet_Deals.write(f'E{m}', Deal["Login"])
            worksheet_Deals.write(f'F{m}', Deal["Currency"])
            worksheet_Deals.write(f'G{m}', Deal["Deal"])
            worksheet_Deals.write(f'H{m}', Deal["Order"])
            worksheet_Deals.write(f'I{m}', str(Deal["Time"]))
            worksheet_Deals.write(f'J{m}', Deal["Action"])
            worksheet_Deals.write(f'K{m}', Deal["Entry"])
            worksheet_Deals.write(f'L{m}', Deal["Symbol"])
            worksheet_Deals.write(f'M{m}', Deal["Volume"], number)
            worksheet_Deals.write(f'N{m}', Deal["Volume_USD"], number)
            worksheet_Deals.write(f'O{m}', Deal["Price"],rate)
            worksheet_Deals.write(f'P{m}', Deal["Reason"])
            worksheet_Deals.write(f'Q{m}', Deal["Dealer"])
            worksheet_Deals.write(f'R{m}', Deal["Profit"], number)
            worksheet_Deals.write(f'S{m}', Deal["CBR"],rate)
            worksheet_Deals.write(f'T{m}', Deal["Result"], number)
    my_connection.close()
    workbook_.close()
    wb = xl.Workbooks.Open(direction+agent+" 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx")
    wb.Close(True)
    to_return = {}
    to_return["conv_count"] = str(len(convertations))
    to_return["acc_count"] = str(j-1)
    to_return["oper_count"] = str(m-1)
    to_return["reward_count"] = str(len(login_list))
    return to_return
