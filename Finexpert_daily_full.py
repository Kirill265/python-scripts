import requests
import sys
import os
import calendar
import datetime
from datetime import timedelta
import xlsxwriter
import pymysql
import psycopg2
import openpyxl
from psycopg2.extras import DictCursor
from pymysql.cursors import DictCursor
from TeamWox import TW_text_file
import time
from win32com import client
import win32com
import pandas as pd

def telegram_bot(Report: str):
    api_token = '1362203438:AAFNp5tXRWi6Pn5RkIgqq_7ELHdGTbY9CUs'
    requests.get('https://api.telegram.org/bot{}/sendMessage'.format(api_token), params=dict(
        chat_id='-1001156138635',
        parse_mode= 'Markdown',
        text=Report 
))
    
my_username = 'kcherkasov'
my_password = '6ne6H7O3ikVUvmDc570AMfmIgTSXZkcOI'
my_connection = pymysql.connect(
    host='172.16.1.42',
    port=3307,
    user=my_username,
    password=my_password,
    db='my',
    charset='utf8mb4',
    cursorclass=DictCursor
)
postgre_username = 'kcherkasov'
postgre_password = 'fg8GerFulLmDdWw4PhDjy4u5iIDLc7mW7rTUJBRNkCDcSGCs'
Postgre_connection = psycopg2.connect(
    host='172.16.1.42',
    port=5433,
    user=postgre_username,
    password=postgre_password,
    dbname='finance'
)
Postgre_connection_2 = psycopg2.connect(
    host='172.16.1.42',
    port=5433,
    user=postgre_username,
    password=postgre_password,
    dbname='mt5_report'
)
sources_utm = """
            'finexpert','finexpert005',
            'orenburg001',
            'sterlitamak001',
            'smolensk','smolensk002','smolensk003',
            'bryansk','bryansk002','bryansk003',
            'Курск','kursk','kursk002','kursk003',
            'belgorod','belgorod002','belgorod003',
            'rostov','rostov002','rostov003',
            'voronezh','voronezh002','voronezh002','voronezh003',
            'kazan','kazan002','kazan003',
            'ulianovsk','ulianovsk002','ulianovsk003',
            'toliatti','toliatti002','toliatti003',
            'samara','samara002','samara003',
            'ekaterinburg','ekaterinburg002','ekaterinburg003',
            'norilsk','norilsk002','norilsk003',
            'novosibirsk','novosibirsk002','novosibirsk003',
            'krasnoiarsk','krasnoiarsk002','krasnoiarsk003',
            'Ufa','Ufa001','Ufa002','Ufa003',
            'ufa004','ufa005','ufa006','ufa007','ufa008',
            'pt001','pt002','pt003','pt004','pt005',
            'Vil','Ya001'
"""
month_number_dict = {"1":'января',"2":'февраля',"3":'марта',"4":'апреля',"5":'мая',"6":'июня',"7":'июля',"8":'августа',"9":'сентября',"10":'октября',"11":'ноября',"12":'декабря'} 
now = datetime.datetime.now()
wday = calendar.weekday(now.year, now.month, now.day)
if wday in [0,1,2,3,4,5]:
    report_date = now - timedelta(days=1)
else:
    sys.exit()
month = month_number_dict[str(report_date.month)]
if report_date.month < 10:
    sql_month = '0'+str(report_date.month)
else:
    sql_month = str(report_date.month)
if report_date.day < 10:
    msg_to_day = '0'+str(report_date.day)
else:
    msg_to_day = str(report_date.day)
date_from = str(report_date.year)+'-'+sql_month+'-01 00:00:00'
date_to = str(report_date.year)+'-'+sql_month+'-'+msg_to_day+' 23:59:59'
direction = 'C:/Users/Kirill_Cherkasov/Documents/Reports/Finexpert_weekly/full/'
#os.mkdir(direction+'01-'+str(report_date.day)+' '+month+' '+str(report_date.year))
#direction += '01-'+str(report_date.day)+' '+month+' '+str(report_date.year)+'/'
workbook_ = xlsxwriter.Workbook(direction+'Finexpert 01-'+msg_to_day+' '+month+' '+str(report_date.year)+'.xlsx')
workbook_.formats[0].set_font_size(8.5)
workbook_.formats[0].set_font_name('Tahoma')
bold_blue_ = workbook_.add_format({'bold': True, 'fg_color': '#DDEBF7', 'align': 'center','valign': 'vcenter'})
number = workbook_.add_format({'num_format': '0.00','align': 'right'})
percentage = workbook_.add_format({'num_format': '0.00%','align': 'center','valign': 'vcenter'})
percentage.set_font_size(8.5)
percentage.set_font_name('Tahoma')
rub_reward = workbook_.add_format({'num_format': '0.00" ₽"'})
usd_volume = workbook_.add_format({'num_format': '"$ "0.00'})
worksheet_PL = workbook_.add_worksheet('PL')
#worksheet_PL.set_default_row(10.5)
worksheet_PL.set_row(0, 15)
worksheet_PL.write('A1', 'Login', bold_blue_)
worksheet_PL.write('B1', 'Created', bold_blue_)
worksheet_PL.write('C1', 'FIO', bold_blue_)
worksheet_PL.write('D1', 'UTM', bold_blue_)
worksheet_PL.write('E1', 'LK', bold_blue_)
worksheet_PL.write('F1', 'Currency', bold_blue_)
worksheet_PL.write('G1', 'Deposit', bold_blue_)
worksheet_PL.write('H1', 'Withdrawal', bold_blue_)
worksheet_PL.write('I1', 'Volume Lots', bold_blue_)
worksheet_PL.set_column(0, 8, 15)
worksheet_PL.write('J1', 'Convertation out', bold_blue_)
worksheet_PL.write('K1', 'Convertation in', bold_blue_)
worksheet_PL.set_column(9, 10, 18)
worksheet_PL.write('L1', 'Equity '+str(report_date.day)+'.'+sql_month, bold_blue_)
worksheet_PL.write('M1', 'Balance 1.'+sql_month, bold_blue_)
worksheet_PL.write('N1', 'Balance '+str(report_date.day)+'.'+sql_month, bold_blue_)
worksheet_PL.set_column(11, 13, 15)
worksheet_PL.write('O1', 'P/L (+Commission)', bold_blue_)
worksheet_PL.set_column(14, 14, 20)
worksheet_Reward = workbook_.add_worksheet('Reward')
#worksheet_Reward.set_default_row(10.5)
worksheet_Reward.set_row(0, 15)
worksheet_Reward.write('A1', 'Login', bold_blue_)
worksheet_Reward.write('B1', 'FIO', bold_blue_)
worksheet_Reward.write('C1', 'UTM', bold_blue_)
worksheet_Reward.write('D1', 'LK', bold_blue_)
worksheet_Reward.write('E1', 'Volume Lots', bold_blue_)
worksheet_Reward.write('F1', 'Volume USD', bold_blue_)
worksheet_Reward.set_column(0, 5, 15)
worksheet_Reward.write('G1', 'Вознаграждение, Руб', bold_blue_)
worksheet_Reward.set_column(6, 6, 25)
worksheet_Deals = workbook_.add_worksheet('Deals')
#worksheet_Deals.set_default_row(10.5)
worksheet_Deals.set_row(0, 15)
worksheet_Deals.write('A1', 'Deal', bold_blue_)
worksheet_Deals.set_column(0, 0, 12)
worksheet_Deals.write('B1', 'Order', bold_blue_)
worksheet_Deals.set_column(1, 1, 15)
worksheet_Deals.write('C1', 'Login', bold_blue_)
worksheet_Deals.write('D1', 'UTM', bold_blue_)
worksheet_Deals.set_column(2, 3, 12)
worksheet_Deals.write('E1', 'Time', bold_blue_)
worksheet_Deals.set_column(4, 4, 20)
worksheet_Deals.write('F1', 'Type', bold_blue_)
worksheet_Deals.write('G1', 'Entry', bold_blue_)
worksheet_Deals.set_column(5, 6, 8.5)
worksheet_Deals.write('H1', 'Symbol', bold_blue_)
worksheet_Deals.set_column(7, 7, 13)
worksheet_Deals.write('I1', 'Volume', bold_blue_)
worksheet_Deals.set_column(8, 8, 8.5)
worksheet_Deals.write('J1', 'Volume USD', bold_blue_)
worksheet_Deals.set_column(9, 9, 13)
worksheet_Deals.write('K1', 'Price', bold_blue_)
worksheet_Deals.set_column(10, 10, 8.5)
worksheet_Deals.write('L1', 'Reason', bold_blue_)
worksheet_Deals.set_column(11, 11, 13)
worksheet_Deals.write('M1', 'Profit', bold_blue_)
worksheet_Deals.set_column(12, 12, 8.5)
worksheet_Deals.write('N1', 'Dealer', bold_blue_)
worksheet_Deals.set_column(13, 13, 13)
worksheet_Deals.write('O1', 'Currency', bold_blue_)
worksheet_Deals.set_column(14, 14, 8.5)
worksheet_ML = workbook_.add_worksheet('Margin Level')
#worksheet_ML.set_default_row(10.5)
worksheet_ML.set_row(0, 15)
worksheet_ML.write('A1', 'Login', bold_blue_)
worksheet_ML.write('B1', 'UTM', bold_blue_)
worksheet_ML.write('C1', 'LK', bold_blue_)
worksheet_ML.write('D1', 'Currency', bold_blue_)
worksheet_ML.set_column(0, 3, 15)
worksheet_ML.write('E1', 'Margin Level '+str(report_date.day)+'.'+sql_month, bold_blue_)
worksheet_ML.set_column(4, 4, 20)
workbook_sum = xlsxwriter.Workbook(direction+'finexpert рассчёт 01-'+msg_to_day+' '+month+' '+str(report_date.year)+'.xlsx')
border_sum = workbook_sum.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
rub = workbook_sum.add_format({'num_format': '0.00"₽"'})
usd = workbook_sum.add_format({'num_format': '"$"0.00'})
rub_border = workbook_sum.add_format({'num_format': '0.00"₽"','border': 1,'align': 'center','valign': 'vcenter'})
usd_border = workbook_sum.add_format({'num_format': '"$"0.00','border': 1,'align': 'center','valign': 'vcenter'})
bold = workbook_sum.add_format({'bold': True})
bold_border = workbook_sum.add_format({'bold': True,'border': 1,'align': 'center','valign': 'vcenter'})
workbook_sum.formats[0].set_font_size(8.5)
workbook_sum.formats[0].set_font_name('Tahoma')
worksheet_sum = workbook_sum.add_worksheet('Итог')
worksheet_sum.set_default_row(12)
worksheet_sum.set_row(0, 15)
worksheet_sum.write('A1', 'Агент', bold_border)
worksheet_sum.write('B1', 'Клиент', bold_border)
worksheet_sum.set_column(0, 1, 15)
worksheet_sum.write('C1', 'Сумма за клиента', bold_border)
worksheet_sum.set_column(2, 2, 25)
worksheet_sum.write('D1', 'Итоговая сумма', bold_border)
worksheet_sum.set_column(3, 3, 20)
with my_connection.cursor() as cursor:
    query = """
            SELECT 
            CONCAT("'",c.universal_id,"',") AS UID_for_Postgre 
            FROM customer_utm cu
            LEFT JOIN customer c ON cu.customer_id = c.id
            LEFT JOIN utm u ON cu.utm_id = u.id
            WHERE u.utm_source IN ("""+sources_utm+""")
     """
    cursor.execute(query)
    UIDS = cursor.fetchall()
    customer_id = ''
    for UID in UIDS:
        customer_id += UID["UID_for_Postgre"]
    query = """
            SELECT CONCAT(a.login,',') AS login_comma
            FROM account a
            LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
            LEFT JOIN utm u ON cu.utm_id = u.id 
            WHERE u.utm_source IN ("""+sources_utm+""")
            ORDER BY Login
     """
    cursor.execute(query)
    logins_comma = cursor.fetchall()
    login_for_mt5 = ''
    for login_comma in logins_comma:
        login_for_mt5 += login_comma["login_comma"]
    query = """
            SET @@time_zone = \"+3:00\";
    """
    cursor.execute(query)
    query = """
            SELECT DATE(crh.date) as currency_date, crh.value FROM currency_rate_history crh
            WHERE crh.from_id = 1
            AND crh.to_id = 3
            AND crh.date BETWEEN DATE(\""""+date_from+"""\") AND DATE(\""""+date_to+"""\")
            ORDER BY crh.date
    """
    cursor.execute(query)
    currency_rates = cursor.fetchall()
    query = """
            SELECT DISTINCT u.utm_source FROM platform_mt5_deal pmd
            LEFT JOIN account a ON pmd.login = a.login
            LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
            LEFT JOIN utm u ON cu.utm_id = u.id
            WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
            and u.utm_source IN ("""+sources_utm+""")
    """
    cursor.execute(query)
    utm_sources = cursor.fetchall()
    k = 1
    done_counter = 0
    login_list = []
    for utm_source in utm_sources:
        query = """
                SELECT a.login FROM account a
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id
                WHERE u.utm_source = '"""+utm_source["utm_source"]+"""'
                AND a.login IN 
                (SELECT DISTINCT pmd.login FROM platform_mt5_deal pmd
                WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\")
        """
        cursor.execute(query)
        logins = cursor.fetchall()
        i = 1
        for login in logins:
            i += 1
            k += 1
            login_list.append(str(login["login"]))
            worksheet_login = workbook_sum.add_worksheet(str(login["login"]))
            worksheet_login.set_default_row(12)
            worksheet_login.set_row(0, 20)
            worksheet_login.write('A1', 'id', bold)
            worksheet_login.write('B1', 'login', bold)
            worksheet_login.set_column(0, 1, 13)
            worksheet_login.write('C1', 'mt5_order_id', bold)
            worksheet_login.set_column(2, 2, 15)
            worksheet_login.write('D1', 'action', bold)
            worksheet_login.write('E1', 'entry', bold)
            worksheet_login.set_column(3, 4, 7)
            worksheet_login.write('F1', 'symbol', bold)
            worksheet_login.set_column(5, 5, 12)
            worksheet_login.write('G1', 'created_at', bold)
            worksheet_login.write('H1', 'date', bold)
            worksheet_login.set_column(6, 7, 18)
            worksheet_login.write('I1', 'price', bold)
            worksheet_login.set_column(8, 8, 8)
            worksheet_login.write('J1', 'volume', bold)
            worksheet_login.set_column(9, 9, 9)
            worksheet_login.write('K1', 'profit', bold)
            worksheet_login.set_column(10, 10, 10) 
            worksheet_login.write('L1', 'volume_usd', bold)
            worksheet_login.set_column(11, 11, 14)
            worksheet_login.write('M1', 'expert_position_id', bold)
            worksheet_login.set_column(12, 12, 20)
            worksheet_login.write('N1', 'Одна поза', bold)
            worksheet_login.set_column(13, 13, 11)
            worksheet_login.write('O1', 'Длительность позиции', bold)
            worksheet_login.set_column(14, 14, 15)
            worksheet_login.write('P1', 'Короткая', bold)
            worksheet_login.set_column(15, 15, 10)
            worksheet_login.write('Q1', 'Вознаграждение, USD', bold)
            worksheet_login.set_column(16, 16, 18)
            worksheet_login.write('R1', 'Курс USDRUR на дату сделки', bold)
            worksheet_login.set_column(17, 17, 14)
            worksheet_login.write('S1', 'Вознаграждение, RUR', bold)
            worksheet_login.set_column(18, 18, 18)
            worksheet_login.write('V1', 'USD без округления', bold)
            worksheet_login.write('W1', 'RUB без округления', bold)
            worksheet_login.set_column(21, 22, 12)
            worksheet_login.write('U2', 'Вознаграждение без учета коротких сделок')
            worksheet_login.write('U3', 'Вознаграждение с вычетом коротких сделок')
            worksheet_login.set_column(20, 20, 23)
            worksheet_login.write('V2', '=SUM(L:L)/1000000*20', usd)
            worksheet_login.write('V3', '=SUM(Q:Q)', usd)
            worksheet_login.write('W3', '=SUM(S:S)', rub)
            query = """
                    SELECT
                    pmd.id
                    , pmd.login
                    , pmd.mt5_order_id
                    , pmd.action
                    , pmd.entry
                    , pmd.symbol
                    , pmd.created_at
                    , date(pmd.created_at) AS date
                    , pmd.price
                    , pmd.volume
                    , pmd.profit
                    , pmd.volume_usd
                    , pmd.expert_position_id
                    FROM platform_mt5_deal pmd
                    LEFT JOIN account a ON pmd.login = a.login
                    LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                    LEFT JOIN utm u ON cu.utm_id = u.id
                    LEFT JOIN account_group ag ON a.group_id = ag.id
                    WHERE a.login = '"""+str(login["login"])+"""'
                    AND
                    pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                    ORDER BY pmd.id
            """
            cursor.execute(query)
            deals = cursor.fetchall()
            j = 1
            for deal in deals:
                j += 1
                worksheet_login.write(f'A{j}', deal["id"])
                worksheet_login.write(f'B{j}', deal["login"])
                worksheet_login.write(f'C{j}', deal["mt5_order_id"])
                worksheet_login.write(f'D{j}', deal["action"])
                worksheet_login.write(f'E{j}', deal["entry"])
                worksheet_login.write(f'F{j}', deal["symbol"])
                worksheet_login.write(f'G{j}', str(deal["created_at"]))
                worksheet_login.write(f'H{j}', str(deal["date"]))
                worksheet_login.write(f'I{j}', deal["price"])
                worksheet_login.write(f'J{j}', deal["volume"])
                worksheet_login.write(f'K{j}', deal["profit"])
                worksheet_login.write(f'L{j}', deal["volume_usd"])
                worksheet_login.write(f'M{j}', deal["expert_position_id"])
                worksheet_login.write(f'N{j}', '=M'+str(j)+'=M'+str(j-1))
                worksheet_login.write(f'O{j}', '=IF(N'+str(j)+' = TRUE,(G'+str(j)+'-G'+str(j-1)+')*24*60*60,"> 10")')
                worksheet_login.write(f'P{j}', '=IF(O'+str(j)+'=">10",FALSE,AND(O'+str(j)+'<10,E'+str(j-1)+'<>1,E'+str(j)+'<>0))')
                worksheet_login.write(f'Q{j}', '=IF(P'+str(j)+' = TRUE,0,L'+str(j)+'/1000000*20)')
                worksheet_login.write(f'R{j}', '=VLOOKUP(H'+str(j)+',\'Курс ЦБ\'!A:B,2,FALSE)')
                worksheet_login.write(f'S{j}', '=Q'+str(j)+'*R'+str(j)+'')
            worksheet_sum.write(f'B{k}',str(login["login"]), border_sum)
            worksheet_sum.write(f'C{k}','=ROUND(\''+worksheet_login.name+'\'!$W$3,2)', rub_border)
        if i == 2:
            worksheet_sum.write(f'D{k}','=C'+str(k), rub_border)
            worksheet_sum.write(f'A{k}',str(utm_source["utm_source"]), border_sum)
        else:
            worksheet_sum.merge_range(f'D{k-i+2}:D{k}','=SUM(C'+str(k-i+2)+':C'+str(k)+')', rub_border)
            worksheet_sum.merge_range(f'A{k-i+2}:A{k}',str(utm_source["utm_source"]), border_sum)
        done_counter += 1
    worksheet_currency = workbook_sum.add_worksheet('Курс ЦБ')
    worksheet_currency.set_default_row(12)
    i = 0
    for currency_rate in currency_rates:
        i += 1
        worksheet_currency.write(f'A{i}',str(currency_rate["currency_date"]))
        worksheet_currency.set_column(0, 0, 12)
        worksheet_currency.write(f'B{i}',currency_rate["value"])
        worksheet_currency.set_column(1, 1, 8)
with Postgre_connection.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:    
    query = """
            SELECT a.login, 
            COALESCE(conv_from.amount, 0.00) as volume_out
            ,COALESCE(conv_to.amount, 0.00) AS volume_in
            FROM account a
            LEFT JOIN 
            (SELECT
            a_from.login AS login
            , SUM(ROUND(- c.amount_from / 100.0, 2)) AS amount
            FROM convertation c
            LEFT JOIN account a_from ON c.account_id_from = a_from.id
            WHERE c.created_at BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
            AND c.status = 3
            GROUP BY a_from.login
            ) AS conv_from ON conv_from.login = a.login
            LEFT JOIN 
            ( SELECT
            a_to.login AS login
            , SUM(ROUND(c.amount_to / 100.0, 2)) AS amount
            FROM convertation c
            LEFT JOIN account a_to ON c.account_id_to = a_to.id
            WHERE c.created_at BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
            AND c.status = 3
            GROUP BY a_to.login
            ) AS conv_to ON conv_to.login = a.login
            WHERE (conv_from.amount IS NOT NULL
            OR conv_to.amount IS NOT NULL
            )
            AND a.customer_id IN (
            -- id клиентов
            """+customer_id[:-1]+"""
            ) 
            ORDER BY a.login
    """
    cursor.execute(query)
    convertations = cursor.fetchall()
    convertation_dict = {}
    for convertation in convertations:
        convertation_dict[str(convertation["login"])] = {"out":convertation["volume_out"], "in":convertation["volume_in"]}
Postgre_connection.close()
workbook_sum.close()
#log_txt = open(direction+'logErr.txt', 'w')
#log_txt.write('1\n')
xl = win32com.client.DispatchEx('Excel.Application')
#log_txt.write('2\n')
xl.Visible = False
#log_txt.write('3\n')
wb = xl.Workbooks.Open(direction+"finexpert рассчёт 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx")
#wb = xl.Workbooks.Open("C:/Users/Kirill_Cherkasov/Documents/Reports/Finexpert_weekly/full/finexpert рассчёт 01-01 апреля 2021.xlsx")
#wb = xl.Workbooks.Open("C:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\Finexpert_weekly\\full\\finexpert рассчёт 01-01 апреля 2021.xlsx")
#log_txt.write('4\n')
wb.Close(True)
#log_txt.write('5\n')
#log_txt.close()
with Postgre_connection_2.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:    
    query = """
            SELECT mt5a."Login"
            , ROUND(COALESCE(Dps.Deposit,0)::NUMERIC,2) AS Deposit
            , ROUND(COALESCE(-Wth.Withdrawal,0)::NUMERIC,2) AS Withdrawal
            , ROUND(COALESCE(PL.profit,0)::NUMERIC,2) AS Profit
            FROM mt5_accounts mt5a
            LEFT JOIN (
            SELECT "Login", SUM("Profit") AS profit FROM mt5_deals
            WHERE "Action" IN (0,1,7)
            AND "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
            GROUP BY "Login") AS PL ON mt5a."Login" = PL."Login"
            LEFT JOIN (
            SELECT "Login", SUM("Profit") AS Deposit FROM mt5_deals
            WHERE "Action" = 2
            AND  "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
            AND (
            "Comment" = ''
            OR
            "Comment" LIKE '%Deposit%'
            OR
            "Comment" LIKE '%Возврат%'
            OR
            "Comment" LIKE '%Refund%'
            )
            GROUP BY "Login") AS Dps ON Dps."Login" = mt5a."Login"
            LEFT JOIN (
            SELECT "Login", SUM("Profit") AS Withdrawal FROM mt5_deals
            WHERE "Action" = 2
            AND  "TimeMsc" BETWEEN \'"""+date_from+"""\' AND \'"""+date_to+"""\'
            AND (
            "Comment" LIKE '%Withdrawal%'
            OR
            "Comment" LIKE '%Удержание%'
            OR
            "Comment" LIKE '%удержание%'
            )
            GROUP BY "Login") AS Wth ON Wth."Login" = mt5a."Login" 
            WHERE mt5a."Login" IN (
            """+login_for_mt5[:-1]+"""
            );
    """
    cursor.execute(query)
    OPRDS_plus_PL = cursor.fetchall()
    OPRDS_PL_dict = {}
    for OPRDS_PL in OPRDS_plus_PL:
        OPRDS_PL_dict[str(OPRDS_PL["Login"])] = {"deposit":OPRDS_PL["deposit"], "withdrawal":OPRDS_PL["withdrawal"], "profit":OPRDS_PL["profit"]}
Postgre_connection_2.close()
with my_connection.cursor() as cursor:
    query = """
            SET @@time_zone = "+3:00";
     """
    cursor.execute(query)
    query = """
            SELECT a.login AS 'Login'
            , a.created_at AS 'Create'
            , CONCAT(ci.last_name_ru,' ',SUBSTRING(ci.first_name_ru,1,1),'.',SUBSTRING(ci.middle_name_ru,1,1),'.') AS 'FIO'
            , u.utm_source AS 'UTM'
            , a.customer_id AS 'LK'
            , c.name AS 'Currency'
            , IFNULL(ROUND(pmd.volume_sum, 2), 0) AS 'Volume_Lots'
            , IFNULL(ROUND(EquityTo.Equity,2),0) AS Equity_to
            , IFNULL(ROUND(EquityFrom.Balance,2),0) AS Balance_from
            , IFNULL(ROUND(EquityTo.Balance,2),0) AS Balance_to
            , IFNULL(ROUND(EquityTo.MarginLevel/100,4),0) AS Margin_Level
            FROM account a
            LEFT JOIN currency c ON a.currency_id = c.id
            LEFT JOIN customer_individual ci ON a.customer_id = ci.customer_id
            LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
            LEFT JOIN utm u ON cu.utm_id = u.id 
            LEFT JOIN
            (
            SELECT platform_mt5_deal.login
            , SUM(platform_mt5_deal.volume) AS volume_sum
            FROM platform_mt5_deal 
            WHERE platform_mt5_deal.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
            GROUP BY platform_mt5_deal.login
            ) AS pmd ON pmd.login = a.login
              LEFT JOIN
            (
            SELECT
            m5ad1.login
            , m5ad1.MarginLevel
            , m5ad1.Balance
            , m5ad1.Equity
            FROM mt5_events.mt5accountdaily m5ad1
            WHERE DATE(DATE_ADD("1970-01-01 00:00:00", INTERVAL ROUND(m5ad1.TimestampMS/1000 - 3600, 0) SECOND)) = DATE(\""""+date_to+"""\")
            ) AS EquityTo ON a.login = EquityTo.login
            LEFT JOIN
            (
            SELECT
            m5ad2.login
            , m5ad2.Balance
            FROM mt5_events.mt5accountdaily m5ad2
            WHERE DATE(DATE_ADD("1970-01-01 00:00:00", INTERVAL ROUND(m5ad2.TimestampMS/1000 - 3600, 0) SECOND)) = DATE(DATE_ADD(\""""+date_from+"""\", INTERVAL -1 DAY))
            ) AS EquityFrom ON a.login = EquityFrom.login
            WHERE u.utm_source IN ("""+sources_utm+""")
            ORDER BY Login
     """
    cursor.execute(query)
    PL_all = cursor.fetchall()
    wb = openpyxl.load_workbook(direction+"finexpert рассчёт 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx",data_only=True)
    j = 1
    for PL_one in PL_all:
        j += 1
        worksheet_PL.write(f'A{j}', PL_one["Login"])
        worksheet_PL.write(f'B{j}', str(PL_one["Create"]))
        worksheet_PL.write(f'C{j}', PL_one["FIO"])
        worksheet_PL.write(f'D{j}', PL_one["UTM"])
        worksheet_PL.write(f'E{j}', PL_one["LK"])
        worksheet_PL.write(f'F{j}', PL_one["Currency"])
        worksheet_PL.write(f'G{j}', OPRDS_PL_dict[str(PL_one["Login"])]["deposit"], number)
        worksheet_PL.write(f'H{j}', OPRDS_PL_dict[str(PL_one["Login"])]["withdrawal"], number)
        worksheet_PL.write(f'I{j}', PL_one["Volume_Lots"], number)
        try:
            worksheet_PL.write(f'J{j}', convertation_dict[str(PL_one["Login"])]["out"], number)
            worksheet_PL.write(f'K{j}', convertation_dict[str(PL_one["Login"])]["in"], number)
        except KeyError:
            worksheet_PL.write(f'J{j}', 0.00, number)
            worksheet_PL.write(f'K{j}', 0.00, number)
        worksheet_PL.write(f'L{j}', PL_one["Equity_to"], number)
        worksheet_PL.write(f'M{j}', PL_one["Balance_from"], number)
        worksheet_PL.write(f'N{j}', PL_one["Balance_to"], number)
        worksheet_PL.write(f'O{j}', OPRDS_PL_dict[str(PL_one["Login"])]["profit"], number)
        worksheet_Reward.write(f'A{j}', PL_one["Login"])
        worksheet_Reward.write(f'B{j}', PL_one["FIO"])
        worksheet_Reward.write(f'C{j}', PL_one["UTM"])
        worksheet_Reward.write(f'D{j}', PL_one["LK"])
        worksheet_Reward.write(f'E{j}', PL_one["Volume_Lots"], number)
        if str(PL_one["Login"]) in login_list:
            sheet = wb[str(PL_one["Login"])]
            RewardSelected = sheet.cell(row = 3, column = 22).value
            print(RewardSelected)
            in_USD = round(RewardSelected*1000000/20,2)
            worksheet_Reward.write(f'F{j}', '='+str(in_USD), usd_volume)
            RewardSelected = sheet.cell(row = 3, column = 23).value
            in_RUB = round(RewardSelected,2)
            worksheet_Reward.write(f'G{j}', '='+str(in_RUB), rub_reward)
        else:
            worksheet_Reward.write(f'F{j}', 0.00, usd_volume)
            worksheet_Reward.write(f'G{j}', 0.00, rub_reward)
        worksheet_ML.write(f'A{j}', PL_one["Login"])
        worksheet_ML.write(f'B{j}', PL_one["UTM"])
        worksheet_ML.write(f'C{j}', PL_one["LK"])
        worksheet_ML.write(f'D{j}', PL_one["Currency"])
        worksheet_ML.write(f'E{j}', PL_one["Margin_Level"],percentage)
    query = """
            SELECT c.created_at AS 'Create'
            , CONCAT(ci.last_name_ru,' ',SUBSTRING(ci.first_name_ru,1,1),'.',SUBSTRING(ci.middle_name_ru,1,1),'.') AS 'FIO'
            , u.utm_source AS 'UTM'
            , c.id AS 'LK'
            FROM customer_utm cu
            LEFT JOIN utm u ON cu.utm_id = u.id
            LEFT JOIN customer c ON cu.customer_id = c.id
            LEFT JOIN customer_individual ci ON c.id = ci.customer_id
            WHERE u.utm_source IN ("""+sources_utm+""")
            AND c.id NOT IN (
            SELECT distinct a.customer_id FROM account a
            )
     """
    cursor.execute(query)
    customerID_all = cursor.fetchall()
    for customerID in customerID_all:
        j += 1
        worksheet_PL.write(f'A{j}', '-')
        worksheet_PL.write(f'B{j}', str(customerID["Create"]))
        worksheet_PL.write(f'C{j}', customerID["FIO"])
        worksheet_PL.write(f'D{j}', customerID["UTM"])
        worksheet_PL.write(f'E{j}', customerID["LK"])
        worksheet_PL.write(f'F{j}', '-')
        worksheet_PL.write(f'G{j}', '-')
        worksheet_PL.write(f'H{j}', '-')
        worksheet_PL.write(f'I{j}', '-')
        worksheet_PL.write(f'J{j}', '-')
        worksheet_PL.write(f'K{j}', '-')
        worksheet_PL.write(f'L{j}', '-')
        worksheet_PL.write(f'M{j}', '-')
        worksheet_PL.write(f'N{j}', '-')
        worksheet_PL.write(f'O{j}', '-')
        worksheet_Reward.write(f'A{j}', '-')
        worksheet_Reward.write(f'B{j}', customerID["FIO"])
        worksheet_Reward.write(f'C{j}', customerID["UTM"])
        worksheet_Reward.write(f'D{j}', customerID["LK"])
        worksheet_Reward.write(f'E{j}', '-')
        worksheet_Reward.write(f'F{j}', '-')
        worksheet_Reward.write(f'G{j}', '-')
        worksheet_ML.write(f'A{j}', '-')
        worksheet_ML.write(f'B{j}', customerID["UTM"])
        worksheet_ML.write(f'C{j}', customerID["LK"])
        worksheet_ML.write(f'D{j}', '-')
        worksheet_ML.write(f'E{j}', '-')
    query = """
            SELECT
            pmd.id AS 'Deal'
            , pmd.mt5_order_id AS 'Order'
            , pmd.login AS 'Login'
            , u.utm_source AS 'UTM'
            , pmd.created_at AS 'Time'
            , IF(pmd.action = 0, 'buy', 'sell') AS 'Type'
            , IF(pmd.entry = 0, 'in', IF(pmd.entry = 1, 'out', 'in/out')) AS 'Entry'
            , pmd.symbol AS 'Symbol'
            , pmd.volume AS 'Volume'
            , pmd.volume_usd AS 'Volume_USD'
            , pmd.price AS 'Price'
            , CASE
            WHEN pmd.reason = 0 THEN 'Client'
            WHEN pmd.reason = 1 THEN 'Expert'
            WHEN pmd.reason = 2 THEN 'Dealer'
            WHEN pmd.reason = 3 THEN 'Stop loss'
            WHEN pmd.reason = 4 THEN 'Take profit'
            WHEN pmd.reason = 5 THEN 'Stop out'
            WHEN pmd.reason = 16 THEN 'Mobile'
            WHEN pmd.reason = 17 THEN 'Web'
            ELSE pmd.reason
            END AS 'Reason'
            , pmd.profit AS 'Profit'
            , pmd.dealer_id AS 'Dealer'
            , UPPER(c.name) AS 'Currency'
            FROM platform_mt5_deal pmd
            LEFT JOIN account a ON pmd.login = a.login
            LEFT JOIN currency c ON a.currency_id = c.id
            LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
            LEFT JOIN utm u ON cu.utm_id = u.id
            WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
            AND (
            u.utm_source IN ("""+sources_utm+""")
            )
            ORDER BY Deal
    """
    cursor.execute(query)
    Deals = cursor.fetchall()
    m = 1
    for Deal in Deals:
        m += 1
        worksheet_Deals.write(f'A{m}', Deal["Deal"])
        worksheet_Deals.write(f'B{m}', Deal["Order"])
        worksheet_Deals.write(f'C{m}', Deal["Login"])
        worksheet_Deals.write(f'D{m}', Deal["UTM"])
        worksheet_Deals.write(f'E{m}', str(Deal["Time"]))
        worksheet_Deals.write(f'F{m}', Deal["Type"])
        worksheet_Deals.write(f'G{m}', Deal["Entry"])
        worksheet_Deals.write(f'H{m}', Deal["Symbol"])
        worksheet_Deals.write(f'I{m}', Deal["Volume"], number)
        worksheet_Deals.write(f'J{m}', Deal["Volume_USD"], number)
        worksheet_Deals.write(f'K{m}', Deal["Price"])
        worksheet_Deals.write(f'L{m}', Deal["Reason"])
        worksheet_Deals.write(f'M{m}', Deal["Profit"], number)
        worksheet_Deals.write(f'N{m}', Deal["Dealer"])
        worksheet_Deals.write(f'O{m}', Deal["Currency"])
my_connection.close()
workbook_.close()

Report_finexpert = """[Отчет по Finexpert](https://team.alfaforex.com/servicedesk/view/11492)

Отчетная дата: *"""+msg_to_day+""" """+month+""" """+str(report_date.year)+"""*.

Счетов: *"""+str(j-1)+"""*
Торговых операций: *"""+str(m-1)+"""*
Конвертирующих счетов: *"""+str(len(convertations))+"""*
Вознаграждение за *"""+str(len(login_list))+"""* счетов."""

telegram_bot(Report_finexpert)
#print(Report_finexpert)

URL_TW = "https://team.alfaforex.com/servicedesk/view/11492"
message_text = ''
attached_file = "C:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\Finexpert_weekly\\full\\"+"Finexpert 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx"+"\nC:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\Finexpert_weekly\\full\\"+"finexpert рассчёт 01-"+msg_to_day+" "+month+" "+str(report_date.year)+".xlsx"

TW_text_file(URL_TW,message_text,attached_file)
