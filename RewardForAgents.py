import requests
import sys
import os
import pymysql
from pymysql.cursors import DictCursor
import xlsxwriter
import datetime
#import calendar
from datetime import timedelta
from TeamWox import TW_text_file
import time
import openpyxl
from win32com import client
import win32com
'''
from PyQt5.QtGui     import *
from PyQt5.QtCore    import *
from PyQt5.QtWidgets import *

class Form(QMainWindow):
    def direct(self):
        return QFileDialog.getExistingDirectory(self,"Укажите путь для сохранения папки с отчетами","")

    def login(self, hostname, port):
        text, ok = QInputDialog.getText(self, hostname+':'+str(port),'login:')
        if ok:
            return text
        else:
            return exit()
    
    def password(self, username):
        text, ok = QInputDialog.getText(self, username,'password:')
        if ok:
            return text
        else:
            return 'Pass'
        
    def month(self):
        text, ok = QInputDialog.getText(self, 'Введите номер отчетного месяца, ','Номер месяца:')
        if ok:
            return text
        else:
            return exit()
'''

def telegram_bot(Report: str):
    api_token = '1362203438:AAFNp5tXRWi6Pn5RkIgqq_7ELHdGTbY9CUs'
    requests.get('https://api.telegram.org/bot{}/sendMessage'.format(api_token), params=dict(
        chat_id='-1001156138635',
        parse_mode= 'Markdown',
        text=Report 
))

hostname='172.16.1.42'
portnum = 3307
'''
app = QApplication(sys.argv)
explorer = Form()
flag = False
while flag == False:
    username = explorer.login(hostname, portnum)
    password = explorer.password(username)
    if password != 'Pass':
        flag = True
username = input('login:')
print('login: ',username,'\n')
password = input('password: ')
'''
username = 'kcherkasov'
password = '6ne6H7O3ikVUvmDc570AMfmIgTSXZkcOI'
connection = pymysql.connect(
    host=hostname,
    port=portnum,
    user=username,
    password=password,
    db='my',
    charset='utf8mb4',
    cursorclass=DictCursor
)
#month = input('Введите месяц для расчета вознаграждения:\t'
month_number_dict = {"1":'январь',"2":'февраль',"3":'март',"4":'апрель',"5":'май',"6":'июнь',"7":'июль',"8":'август',"9":'сентябрь',"10":'октябрь',"11":'ноябрь',"12":'декабрь'}
'''
flag = False
while flag == False:
    month_number = explorer.month()
    if month_number[0] == '0':
        month_number = month_number[-1]
    try:
        month = month_number_dict[month_number]
        flag = True
    except KeyError:
        flag = False
        print('Попробуйте еще раз.')
month = 'октябрь'
'''
now = datetime.datetime.now()
report_date = now - timedelta(days=now.day)
month = month_number_dict[str(report_date.month)]
#month_from = {"январь":'01',"февраль":'02',"март":'03',"апрель":'04',"май":'05',"июнь":'06',"июль":'07',"август":'08',"сентябрь":'09',"октябрь":'10',"ноябрь":'11',"декабрь":'12'}
#month_to = {"январь":'01-31',"февраль":'02-28',"март":'03-31',"апрель":'04-30',"май":'05-31',"июнь":'06-30',"июль":'07-31',"август":'08-31',"сентябрь":'09-30',"октябрь":'10-31',"ноябрь":'11-30',"декабрь":'12-31'}
if report_date.month < 10:
    sql_month = '0'+str(report_date.month)
else:
    sql_month = str(report_date.month)
date_from = str(report_date.year)+'-'+sql_month+'-01 00.00.00'
date_to = str(report_date.year)+'-'+sql_month+'-'+str(report_date.day)+' 23.59.59'
#utm_dict = {" 211504": '14', "6": '20', "212485": '14', "212317": '14', "212544": '14', "212477": '14', "7af": '14', "200343": '14', "201340": '14', "10af": '25', "213330": '20', "13": '14', "212873": '14', "12": '20', "15": '14', "214168": '14', "213889": '14', "213894": '14', "200046": '20', "21": '20', "22": '20', "23": '20', "25af": '14', "215029af": '14', "200336af": '14', "30af": '14', "31af": '20', "32af": '14', "33af": '20',  "35af": '14', "201505af": '14', "36af": '14', "38af": '14', "39af": '20', "217518af": '14', "201206af": '14', "43af": '20', "220131af": '14', "218525af": '14', "46af": '14', "220161af": '20', "212615af": '20', "214725af": '20', "50af": '14', "52af": '14', "51af": '20', "222265af": '14', "222493af": '20', "221708af": '14', "55af": '20', "222362af": '14', "56af": '20', "211459af": '14', "224231af": '14', "61af": '20', "224254af": '14', "63af": '20', "224797af": '14', "224824af": '14', "224066af": '14', "67af": '14', "68af": '20', "68a": '20', "69af": '20', "70af": '14', "214763af": '20', "72af": '20', "216520af": '14', "227460af": '14', "74af": '20'}
utm_txt = open(u'C:/Users/Kirill_Cherkasov/Documents/Python_scripts/utm_percentage.txt', 'r')
utm_dict = {}
for percent in utm_txt:
    utm_dict[percent.split(':')[0]] = percent.split(':')[1].split('\n')[0]
utm_txt.close()
#direction = explorer.direct()+'/'
#direction = 'C:/Users/Kirill_Cherkasov/Documents/Reports/'
direction = 'C:/Users/Kirill_Cherkasov/Documents/Reports/Agents rewards/'
os.mkdir(direction+str(report_date.year)+' '+month)
log_txt = open(direction+str(report_date.year)+' '+month+'/log.txt', 'w')
log_txt.write('пользователь: '+username+'\n')
log_txt.write('месяц: '+month+'\n')
log_txt.write('начало расчета: '+str(datetime.datetime.now())+'\n')
Report_success = ''
Report_unsuccess = '   '
workbook_sum = xlsxwriter.Workbook(direction+str(report_date.year)+' '+month+'/Суммарное вознаграждение за '+month+' '+str(report_date.year)+'.xlsx')
rub_sum = workbook_sum.add_format({'num_format': '0.00"₽"','border': 1,'align': 'center','valign': 'vcenter'})
border_sum = workbook_sum.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
bold_sum = workbook_sum.add_format({'bold': True,'border': 1,'align': 'center','valign': 'vcenter'})
workbook_sum.formats[0].set_font_size(8.5)
workbook_sum.formats[0].set_font_name('Tahoma')
worksheet_sum = workbook_sum.add_worksheet()
worksheet_sum.set_default_row(12)
worksheet_sum.set_row(0, 15)
worksheet_sum.write('A1', 'Агент', bold_sum)
worksheet_sum.set_column(0, 0, 12)
worksheet_sum.write('B1', 'Клиент', bold_sum)
worksheet_sum.set_column(1, 1, 15)
worksheet_sum.write('C1', 'Сумма за клиента', bold_sum)
worksheet_sum.set_column(2, 2, 25)
worksheet_sum.write('D1', 'Итоговая сумма', bold_sum)
worksheet_sum.set_column(3, 3, 20)
with connection.cursor() as cursor:
    query = """
            SET @@time_zone = \"+3:00\";
    """
    cursor.execute(query)
    query = """
            SELECT DATE(crh.date) as currency_date, crh.value FROM currency_rate_history crh
            WHERE crh.from_id = 1
            AND crh.to_id = 3
            AND crh.date BETWEEN DATE(\""""+date_from+"""\") AND DATE(\""""+date_to+"""\")
            ORDER BY crh.date
    """
    cursor.execute(query)
    currency_rates = cursor.fetchall()
    query = """
            SELECT DISTINCT u.utm_source FROM platform_mt5_deal pmd
            LEFT JOIN account a ON pmd.login = a.login
            LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
            LEFT JOIN utm u ON cu.utm_id = u.id
            
            WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
            and u.utm_source NOT IN ('yandex','google','forex','alfaforex','Forex_Ratings','sravniru',
            'desktop','recketmc_mobile','Forex-ratings','QuietMedia','OTM','MTS','raitingfx',
            'yandex_bayan','cpaexchange','beeline','forexru','null','ihodlcom_brokers','rbc_mobile','email','traders-union','infoclub','rns','bank','bank_form','email_mt4web')
            AND u.utm_source NOT IN ('24','24a','24b','26a','fxclub','10af')
            AND u.utm_source NOT IN (
            'finexpert','finexpert005',
            'orenburg001',
            'sterlitamak001',
            'smolensk','smolensk002','smolensk003',
            'bryansk','bryansk002','bryansk003',
            'Курск','kursk','kursk002','kursk003',
            'belgorod','belgorod002','belgorod003',
            'rostov','rostov002','rostov003',
            'voronezh002','voronezh002','voronezh003',
            'kazan','kazan002','kazan003',
            'ulianovsk','ulianovsk002','ulianovsk003',
            'toliatti','toliatti002','toliatti003',
            'samara','samara002','samara003',
            'ekaterinburg','ekaterinburg002','ekaterinburg003',
            'norilsk','norilsk002','norilsk003',
            'novosibirsk','novosibirsk002','novosibirsk003',
            'krasnoiarsk','krasnoiarsk002','krasnoiarsk003',
            'Ufa','Ufa001','Ufa002','Ufa003','Ya001'
            )
            AND u.utm_source IS NOT NULL;
            
            -- WHERE u.utm_source = '214725af'
    """
    cursor.execute(query)
    utm_sources = cursor.fetchall()
    k = 1
    k2 = 1
    done_counter = 0
    attached_utm_files = ""
    for utm_source in utm_sources:
        try:
            utm_dict[utm_source["utm_source"]]
        except KeyError:
            log_txt.write(str(utm_sources.index(utm_source)+1)+') '+utm_source["utm_source"]+': не задан процент вознаграждения\n')
            Report_unsuccess += utm_source["utm_source"]+', '
            continue
        workbook = xlsxwriter.Workbook(direction+str(report_date.year)+' '+month+'/'+utm_source["utm_source"]+' '+month+' '+str(report_date.year)+'.xlsx')
        rub = workbook.add_format({'num_format': '0.00"₽"'})
        usd = workbook.add_format({'num_format': '"$"0.00'})
        rub_border = workbook.add_format({'num_format': '0.00"₽"','border': 1,'align': 'center','valign': 'vcenter'})
        usd_border = workbook.add_format({'num_format': '"$"0.00','border': 1,'align': 'center','valign': 'vcenter'})
        border = workbook.add_format({'border': 1,'align': 'center','valign': 'vcenter'})
        bold = workbook.add_format({'bold': True})
        bold_border = workbook.add_format({'bold': True,'border': 1,'align': 'center','valign': 'vcenter'})
        workbook.formats[0].set_font_size(8.5)
        workbook.formats[0].set_font_name('Tahoma')
        worksheet_itog = workbook.add_worksheet('Итог')
        worksheet_itog.set_default_row(12)
        worksheet_itog.set_row(0, 15)
        worksheet_itog.write('A1', 'Клиент', bold_border)
        worksheet_itog.set_column(0, 0, 12)
        worksheet_itog.write('B1', 'Сумма за клиента RUB', bold_border)
        worksheet_itog.set_column(1, 1, 25)
        worksheet_itog.write('C1', 'Итоговая сумма', bold_border)
        worksheet_itog.set_column(2, 2, 20)
        query = """
                SELECT a.login FROM account a
                LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                LEFT JOIN utm u ON cu.utm_id = u.id
                WHERE u.utm_source = '"""+utm_source["utm_source"]+"""'
                AND a.login IN 
                (SELECT DISTINCT pmd.login FROM platform_mt5_deal pmd
                WHERE pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\")
        """
        cursor.execute(query)
        logins = cursor.fetchall()
        i = 1
        for login in logins:
            i += 1
            k += 1
            worksheet_login = workbook.add_worksheet(str(login["login"]))
            worksheet_login.set_default_row(12)
            worksheet_login.set_row(0, 20)
            worksheet_login.write('A1', 'id', bold)
            worksheet_login.write('B1', 'login', bold)
            worksheet_login.set_column(0, 1, 13)
            worksheet_login.write('C1', 'mt5_order_id', bold)
            worksheet_login.set_column(2, 2, 15)
            worksheet_login.write('D1', 'action', bold)
            worksheet_login.write('E1', 'entry', bold)
            worksheet_login.set_column(3, 4, 7)
            worksheet_login.write('F1', 'symbol', bold)
            worksheet_login.set_column(5, 5, 12)
            worksheet_login.write('G1', 'created_at', bold)
            worksheet_login.write('H1', 'date', bold)
            worksheet_login.set_column(6, 7, 18)
            worksheet_login.write('I1', 'price', bold)
            worksheet_login.set_column(8, 8, 8)
            worksheet_login.write('J1', 'volume', bold)
            worksheet_login.set_column(9, 9, 9)
            worksheet_login.write('K1', 'profit', bold)
            worksheet_login.set_column(10, 10, 10) 
            worksheet_login.write('L1', 'volume_usd', bold)
            worksheet_login.set_column(11, 11, 14)
            worksheet_login.write('M1', 'expert_position_id', bold)
            worksheet_login.set_column(12, 12, 20)
            worksheet_login.write('N1', 'Одна поза', bold)
            worksheet_login.set_column(13, 13, 11)
            worksheet_login.write('O1', 'Длительность позиции', bold)
            worksheet_login.set_column(14, 14, 15)
            worksheet_login.write('P1', 'Короткая', bold)
            worksheet_login.set_column(15, 15, 10)
            worksheet_login.write('Q1', 'Вознаграждение, USD', bold)
            worksheet_login.set_column(16, 16, 18)
            worksheet_login.write('R1', 'Курс USDRUR на дату сделки', bold)
            worksheet_login.set_column(17, 17, 14)
            worksheet_login.write('S1', 'Вознаграждение, RUR', bold)
            worksheet_login.set_column(18, 18, 18)
            worksheet_login.write('V1', 'USD без округления', bold)
            worksheet_login.write('W1', 'RUB без округления', bold)
            worksheet_login.set_column(21, 22, 12)
            worksheet_login.write('U2', 'Вознаграждение без учета коротких сделок')
            worksheet_login.write('U3', 'Вознаграждение с вычетом коротких сделок')
            worksheet_login.set_column(20, 20, 23)
            worksheet_login.write('V2', '=SUM(L:L)/1000000*'+utm_dict[utm_source["utm_source"]], usd)
            worksheet_login.write('V3', '=SUM(Q:Q)', usd)
            worksheet_login.write('W3', '=SUM(S:S)', rub)
            query = """
                    SELECT
                    pmd.id
                    , pmd.login
                    , pmd.mt5_order_id
                    , pmd.action
                    , pmd.entry
                    , pmd.symbol
                    , pmd.created_at
                    , date(pmd.created_at) AS date
                    , pmd.price
                    , pmd.volume
                    , pmd.profit
                    , pmd.volume_usd
                    , pmd.expert_position_id
                    FROM platform_mt5_deal pmd
                    LEFT JOIN account a ON pmd.login = a.login
                    LEFT JOIN customer_utm cu ON a.customer_id = cu.customer_id
                    LEFT JOIN utm u ON cu.utm_id = u.id
                    LEFT JOIN account_group ag ON a.group_id = ag.id
                    WHERE a.login = '"""+str(login["login"])+"""'
                    AND
                    pmd.created_at BETWEEN \""""+date_from+"""\" AND \""""+date_to+"""\"
                    ORDER BY pmd.id
            """
            cursor.execute(query)
            deals = cursor.fetchall()
            j = 1
            for deal in deals:
                j += 1
                worksheet_login.write(f'A{j}', deal["id"])
                worksheet_login.write(f'B{j}', deal["login"])
                worksheet_login.write(f'C{j}', deal["mt5_order_id"])
                worksheet_login.write(f'D{j}', deal["action"])
                worksheet_login.write(f'E{j}', deal["entry"])
                worksheet_login.write(f'F{j}', deal["symbol"])
                worksheet_login.write(f'G{j}', str(deal["created_at"]))
                worksheet_login.write(f'H{j}', str(deal["date"]))
                worksheet_login.write(f'I{j}', deal["price"])
                worksheet_login.write(f'J{j}', deal["volume"])
                worksheet_login.write(f'K{j}', deal["profit"])
                worksheet_login.write(f'L{j}', deal["volume_usd"])
                worksheet_login.write(f'M{j}', deal["expert_position_id"])
                worksheet_login.write(f'N{j}', '=M'+str(j)+'=M'+str(j-1))
                worksheet_login.write(f'O{j}', '=IF(N'+str(j)+' = TRUE,(G'+str(j)+'-G'+str(j-1)+')*24*60*60,"> 10")')
                worksheet_login.write(f'P{j}', '=IF(O'+str(j)+'=">10",FALSE,AND(O'+str(j)+'<10,E'+str(j-1)+'<>1,E'+str(j)+'<>0))')
                worksheet_login.write(f'Q{j}', '=IF(P'+str(j)+' = TRUE,0,L'+str(j)+'/1000000*'+utm_dict[utm_source["utm_source"]]+')')
                worksheet_login.write(f'R{j}', '=VLOOKUP(H'+str(j)+',\'Курс ЦБ\'!A:B,2,FALSE)')
                worksheet_login.write(f'S{j}', '=Q'+str(j)+'*R'+str(j)+'')
            worksheet_itog.write(f'A{i}',str(login["login"]), border)
            worksheet_itog.write(f'B{i}','=ROUND(\''+worksheet_login.name+'\'!$W$3,2)', rub_border)
            #worksheet_sum.write(f'B{k}',str(login["login"]), border_sum)
            #worksheet_sum.write(f'C{k}','=\''+direction+month+'/['+utm_source["utm_source"]+' '+month+' 2020.xlsx]'+worksheet_login.name+'\'!$W$3', rub_sum)
            #worksheet_sum.write(f'C{k}','=\'['+utm_source["utm_source"]+' '+month+' 2020.xlsx]'+worksheet_login.name+'\'!$W$3', rub_sum)
            #worksheet_sum.write(f'C{k}','=\''+direction+month+'/['+utm_source["utm_source"]+' '+month+' 2020.xlsx]Итог\'!$B$'+str(i), rub_sum)
            #worksheet_sum.write(f'C{k}','=\'['+utm_source["utm_source"]+' '+month+' '+str(report_date.year)+'.xlsx]Итог\'!$B$'+str(i), rub_sum)
        if i == 2:
            worksheet_itog.write('C2','=B2', rub_border)
            worksheet_sum.write(f'D{k}','=C'+str(k), rub_sum)
            worksheet_sum.write(f'A{k}',str(utm_source["utm_source"]), border_sum)
        else:
            worksheet_itog.merge_range(f'C2:C{i}','=SUM(B2:B'+str(i)+')', rub_border)
            worksheet_sum.merge_range(f'D{k-i+2}:D{k}','=SUM(C'+str(k-i+2)+':C'+str(k)+')', rub_sum)
            worksheet_sum.merge_range(f'A{k-i+2}:A{k}',str(utm_source["utm_source"]), border_sum)
        worksheet_currency = workbook.add_worksheet('Курс ЦБ')
        worksheet_currency.set_default_row(12)
        i = 0
        for currency_rate in currency_rates:
            i += 1
            worksheet_currency.write(f'A{i}',str(currency_rate["currency_date"]))
            worksheet_currency.set_column(0, 0, 12)
            worksheet_currency.write(f'B{i}',currency_rate["value"])
            worksheet_currency.set_column(1, 1, 8)
        workbook.close()
        xl = win32com.client.DispatchEx('Excel.Application')
        xl.Visible = False
        wb = xl.Workbooks.Open(direction+str(report_date.year)+" "+month+"/"+utm_source["utm_source"]+" "+month+" "+str(report_date.year)+".xlsx")
        wb.Close(True)
        wb = openpyxl.load_workbook(direction+str(report_date.year)+" "+month+"/"+utm_source["utm_source"]+" "+month+" "+str(report_date.year)+".xlsx",data_only=True)
        for login in logins:
            k2 += 1
            sheet = wb[str(login["login"])]
            RewardSelected = sheet.cell(row = 3, column = 23).value
            in_RUB = round(RewardSelected,2)
            worksheet_sum.write(f'B{k2}',str(login["login"]), border_sum)
            worksheet_sum.write(f'C{k2}','='+str(in_RUB), rub_sum)
        done_counter += 1
        attached_utm_files += "\nC:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\Agents rewards\\"+str(report_date.year)+" "+month+"\\"+utm_source["utm_source"]+" "+month+" "+str(report_date.year)+".xlsx"
        #attached_utm_files += "\nC:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\"+str(report_date.year)+" "+month+"\\"+utm_source["utm_source"]+" "+month+" "+str(report_date.year)+".xlsx"
        log_txt.write(str(utm_sources.index(utm_source)+1)+') '+utm_source["utm_source"]+': done (USD '+utm_dict[utm_source["utm_source"]]+' / 1000000)\n')
        Report_success += '   '+utm_source["utm_source"]+' ('+utm_dict[utm_source["utm_source"]]+' USD)\n'
connection.close()
workbook_sum.close()
log_txt.write('рассчитано для '+str(done_counter)+' агентов из '+ str(len(utm_sources))+'\n')
log_txt.write('конец расчета: '+str(datetime.datetime.now()))
log_txt.close()
if done_counter == len(utm_sources):
    Report_unsuccess = ''
else:
    Report_unsuccess = """
Для следующих агентов не задан процент вознаграждения:
*"""+Report_unsuccess[:-2]+"""*
"""
    
Report_reward = """[Расчет вознаграждения для агентов](https://team.alfaforex.com/servicedesk/view/11534)

Отчетный месяц: *"""+month+""" """+str(report_date.year)+"""*.

Рассчитано для *"""+str(done_counter)+""" / """+ str(len(utm_sources))+"""* агентов:
*"""+Report_success+"""*"""+Report_unsuccess

telegram_bot(Report_reward)
#print(Report_reward)

URL_TW = "https://team.alfaforex.com/servicedesk/view/11534"
message_text = 'За '+month+' '+str(report_date.year)
attached_file = "C:\\Users\\Kirill_Cherkasov\\Documents\\Reports\\Agents rewards\\"+str(report_date.year)+" "+month+"\\Суммарное вознаграждение за "+month+" "+str(report_date.year)+".xlsx"+attached_utm_files
TW_text_file(URL_TW,message_text,attached_file)
#print(attached_file)
